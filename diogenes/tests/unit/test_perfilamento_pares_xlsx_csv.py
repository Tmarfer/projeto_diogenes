"""
tests/unit/test_perfilamento_pares_xlsx_csv.py — DVA-CBS | Projeto Diógenes

Testes com 3 pares reais XLSX + CSV, simulando a estrutura do pacote MOD_010:

  Par 1 — Créditos PF
    AUX_MOD_010_creditos_pf.xlsx   (2 abas: Consolidado 2k linhas, Detalhes 800 linhas)
    base_creditos_pf_2023.csv      (2.000 linhas — mesma estrutura da aba Consolidado)
    Cenários: CPF com 80 duplicatas no XLSX; VALOR_CREDITO com 10% nulos no CSV

  Par 2 — Notas Fiscais de Entrada
    AUX_MOD_010_nfe_entrada.xlsx   (1 aba: 4.000 linhas)
    notas_fiscais_entrada_2023.csv (4.000 linhas — espelho do XLSX)
    Cenários: CHAVE_NF única em ambos; CNPJ_EMITENTE com 30% nulos em ambos

  Par 3 — Produtor Rural PF
    AUX_MOD_010_produtor_rural.xlsx (3 abas: Resumo 600 linhas, NCM 300, Parâmetros 50)
    producao_rural_pf_2023.csv      (600 linhas — espelho da aba Resumo)
    Cenários: RECEITA_BRUTA numérica com min/max/média; CPF único em ambos

O que os testes validam:
  - Contagens de linhas exatas (sem truncamento)
  - Detecção de duplicatas em CPF (Par 1 XLSX)
  - Detecção de nulos acima do limiar (CNPJ_EMITENTE Par 2 → alerta ALTA)
  - Stats numéricas min/max/média (RECEITA_BRUTA Par 3)
  - Colunas compartilhadas cross-file (CPF aparece em xlsx + csv do Par 1 e Par 3)
  - 6 arquivos = 6 .perfil.yaml + 1 perfil_pacote.yaml
  - Watson recebe prefixo analítico via file_prep.preparar_arquivo
"""
from __future__ import annotations

import csv
from pathlib import Path

import openpyxl
import pytest

from diogenes.motors.motor_perfilamento import (
    ler_perfil_texto,
    perfilar_arquivo,
    perfilar_diretorio,
)

# ── Geradores de dados ────────────────────────────────────────────────────────

def _gerar_xlsx_creditos(path: Path) -> None:
    """
    AUX_MOD_010_creditos_pf.xlsx
      Aba 'Consolidado': 2.000 linhas, CPF com 80 duplicatas, VALOR_CREDITO, ANO_BASE, NCM
      Aba 'Detalhes':     800 linhas, CPF, DESCRICAO_CREDITO, VALOR_ITEM, DATA_APURACAO
    """
    wb = openpyxl.Workbook()

    ws_c = wb.active
    ws_c.title = "Consolidado"
    ws_c.append(["CPF", "VALOR_CREDITO", "ANO_BASE", "NCM", "TIPO_CONTRIBUINTE"])
    base = 30000000000
    for i in range(1920):          # 1920 únicos
        cpf = str(base + i).zfill(11)
        ws_c.append([cpf, round(200.0 + i * 1.5, 2), 2023, f"22{i % 99:02d}", "PF"])
    for i in range(80):            # 80 duplicatas: repete os 80 primeiros CPFs
        cpf = str(base + i).zfill(11)
        ws_c.append([cpf, round(50.0 + i, 2), 2023, f"22{i:02d}", "PF"])

    ws_d = wb.create_sheet("Detalhes")
    ws_d.append(["CPF", "DESCRICAO_CREDITO", "VALOR_ITEM", "DATA_APURACAO"])
    for i in range(800):
        cpf = str(base + i).zfill(11)
        ws_d.append([cpf, f"CREDITO_ITEM_{i}", round(10.0 + i * 0.2, 2), f"2023-{(i % 12) + 1:02d}-01"])

    wb.save(path)


def _gerar_csv_creditos(path: Path) -> None:
    """
    base_creditos_pf_2023.csv — espelha a aba Consolidado do XLSX
    2.000 linhas; CPF único (sem duplicatas); VALOR_CREDITO com ~10% nulos.
    """
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["CPF", "VALOR_CREDITO", "ANO_BASE", "NCM", "TIPO_CONTRIBUINTE"])
        base = 30000000000
        for i in range(2000):
            cpf = str(base + i).zfill(11)
            valor = "" if i % 10 == 0 else f"{200.0 + i * 1.5:.2f}"   # 10% nulos
            writer.writerow([cpf, valor, 2023, f"22{i % 99:02d}", "PF"])


def _gerar_xlsx_nfe(path: Path) -> None:
    """
    AUX_MOD_010_nfe_entrada.xlsx — 1 aba, 4.000 linhas
    CHAVE_NF única (alfanumérica), CNPJ_EMITENTE com 30% nulos, VALOR_NF numérico.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NFe_Entrada"
    ws.append(["CHAVE_NF", "CNPJ_EMITENTE", "VALOR_NF", "DATA_EMISSAO", "UF_EMITENTE"])
    for i in range(4000):
        chave = f"NF{i:010d}SP{i:010d}CE{i:06d}"   # alfanumérico → não vira HUGEINT
        cnpj = None if i % 3 == 0 else f"{10000000000000 + i:014d}"    # 33% nulos
        ws.append([chave, cnpj, round(100.0 + i * 2.5, 2), f"2023-{(i % 12) + 1:02d}-01", "SP"])
    wb.save(path)


def _gerar_csv_nfe(path: Path) -> None:
    """
    notas_fiscais_entrada_2023.csv — espelha o XLSX, 4.000 linhas
    CHAVE_NF única; CNPJ_EMITENTE com ~30% nulos.
    """
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["CHAVE_NF", "CNPJ_EMITENTE", "VALOR_NF", "DATA_EMISSAO", "UF_EMITENTE"])
        for i in range(4000):
            chave = f"NF{i:010d}SP{i:010d}CE{i:06d}"
            cnpj = "" if i % 3 == 0 else f"{10000000000000 + i:014d}"
            writer.writerow([chave, cnpj, f"{100.0 + i * 2.5:.2f}",
                             f"2023-{(i % 12) + 1:02d}-01", "SP"])


def _gerar_xlsx_produtor_rural(path: Path) -> None:
    """
    AUX_MOD_010_produtor_rural.xlsx — 3 abas
      Resumo:     600 linhas, CPF único, RECEITA_BRUTA, PRODUTO, UF
      NCM:        300 linhas, COD_NCM, DESCRICAO, ALIQUOTA_CBS
      Parâmetros: 50 linhas, PARAMETRO, VALOR_REFERENCIA, JUSTIFICATIVA
    """
    wb = openpyxl.Workbook()

    ws_r = wb.active
    ws_r.title = "Resumo"
    ws_r.append(["CPF", "RECEITA_BRUTA", "PRODUTO", "UF", "ANO_BASE"])
    for i in range(600):
        cpf = str(40000000000 + i).zfill(11)
        receita = round(5000.0 + i * 100.0, 2)   # range: 5.000 a 64.900
        ws_r.append([cpf, receita, f"PRODUTO_{i % 30}", ["SP", "MG", "PR", "RS"][i % 4], 2023])

    ws_n = wb.create_sheet("NCM")
    ws_n.append(["COD_NCM", "DESCRICAO", "ALIQUOTA_CBS"])
    for i in range(300):
        ws_n.append([f"NCM{i:04d}", f"PRODUTO_RURAL_{i}", round(0.04 + i * 0.0005, 6)])

    ws_p = wb.create_sheet("Parâmetros")
    ws_p.append(["PARAMETRO", "VALOR_REFERENCIA", "JUSTIFICATIVA"])
    parametros = [
        ("REDUTOR_PF", 0.20, "Art. 168 §3 LC 214/2025"),
        ("ALIQUOTA_CBS_REF", 0.088, "Acórdão 2833/2025"),
        ("PERCENTUAL_AMOSTRA", 0.05, "Descarte de outliers"),
    ]
    for _ in range(50):
        p, v, j = parametros[_ % len(parametros)]
        ws_p.append([f"{p}_{_:02d}", v, j])

    wb.save(path)


def _gerar_csv_produtor_rural(path: Path) -> None:
    """
    producao_rural_pf_2023.csv — espelha a aba Resumo do XLSX, 600 linhas
    CPF único, RECEITA_BRUTA numérica.
    """
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["CPF", "RECEITA_BRUTA", "PRODUTO", "UF", "ANO_BASE"])
        for i in range(600):
            cpf = str(40000000000 + i).zfill(11)
            writer.writerow([cpf, f"{5000.0 + i * 100.0:.2f}",
                             f"PRODUTO_{i % 30}", ["SP", "MG", "PR", "RS"][i % 4], 2023])


# ── Fixture central: monta o diretório de inputs com os 6 arquivos ────────────

@pytest.fixture(scope="module")
def inputs_dir(tmp_path_factory: pytest.TempPathFactory) -> Path:
    """
    Cria um diretório de inputs com os 3 pares XLSX+CSV.
    scope=module: os arquivos são gerados uma vez e reusados em todos os testes.
    """
    d = tmp_path_factory.mktemp("inputs_pares")

    _gerar_xlsx_creditos(d / "AUX_MOD_010_creditos_pf.xlsx")
    _gerar_csv_creditos(d / "base_creditos_pf_2023.csv")

    _gerar_xlsx_nfe(d / "AUX_MOD_010_nfe_entrada.xlsx")
    _gerar_csv_nfe(d / "notas_fiscais_entrada_2023.csv")

    _gerar_xlsx_produtor_rural(d / "AUX_MOD_010_produtor_rural.xlsx")
    _gerar_csv_produtor_rural(d / "producao_rural_pf_2023.csv")

    return d


@pytest.fixture(scope="module")
def perfis_dir(tmp_path_factory: pytest.TempPathFactory) -> Path:
    return tmp_path_factory.mktemp("perfis")


@pytest.fixture(scope="module")
def pacote(inputs_dir: Path, perfis_dir: Path):
    """Roda perfilar_diretorio uma vez; todos os testes consomem o resultado."""
    return perfilar_diretorio(inputs_dir, perfis_dir)


# ── Testes: cobertura do diretório ────────────────────────────────────────────

class TestCoberturaDiretorio:
    def test_todos_6_arquivos_perfilados(self, pacote) -> None:
        assert pacote.total_arquivos == 6
        assert len(pacote.arquivos_perfilados) == 6
        assert len(pacote.arquivos_com_erro) == 0

    def test_6_perfil_yaml_criados(self, perfis_dir: Path) -> None:
        yamls = list(perfis_dir.glob("*.perfil.yaml"))
        assert len(yamls) == 6

    def test_perfil_pacote_yaml_existe(self, perfis_dir: Path) -> None:
        assert (perfis_dir / "perfil_pacote.yaml").exists()

    def test_zero_erros(self, pacote) -> None:
        assert pacote.arquivos_com_erro == []


# ── Par 1: Créditos PF ────────────────────────────────────────────────────────

class TestParCreditosPF:
    def test_xlsx_row_count_2000_linhas(self, inputs_dir: Path) -> None:
        perfil = perfilar_arquivo(inputs_dir / "AUX_MOD_010_creditos_pf.xlsx")
        assert perfil.row_count == 2000    # aba Consolidado

    def test_xlsx_lista_2_abas(self, inputs_dir: Path) -> None:
        perfil = perfilar_arquivo(inputs_dir / "AUX_MOD_010_creditos_pf.xlsx")
        assert len(perfil.abas) == 2
        assert "Consolidado" in perfil.abas
        assert "Detalhes" in perfil.abas

    def test_xlsx_detecta_80_duplicatas_cpf(self, inputs_dir: Path) -> None:
        perfil = perfilar_arquivo(inputs_dir / "AUX_MOD_010_creditos_pf.xlsx")
        col_cpf = next(c for c in perfil.colunas if c.nome == "CPF")
        assert col_cpf.alerta_duplicatas is True
        assert col_cpf.duplicatas_count >= 80

    def test_xlsx_alerta_critico_por_cpf_duplicado(self, inputs_dir: Path) -> None:
        perfil = perfilar_arquivo(inputs_dir / "AUX_MOD_010_creditos_pf.xlsx")
        criticos = [a for a in perfil.alertas if a.severidade == "CRITICA"]
        assert len(criticos) >= 1

    def test_csv_row_count_2000_linhas(self, inputs_dir: Path) -> None:
        perfil = perfilar_arquivo(inputs_dir / "base_creditos_pf_2023.csv")
        assert perfil.row_count == 2000

    def test_csv_cpf_unico_sem_alerta_duplicata(self, inputs_dir: Path) -> None:
        """No CSV os CPFs são únicos — sem alerta de duplicata."""
        perfil = perfilar_arquivo(inputs_dir / "base_creditos_pf_2023.csv")
        col_cpf = next(c for c in perfil.colunas if c.nome == "CPF")
        assert col_cpf.alerta_duplicatas is False

    def test_csv_valor_credito_10pct_nulos(self, inputs_dir: Path) -> None:
        perfil = perfilar_arquivo(inputs_dir / "base_creditos_pf_2023.csv")
        col_v = next(c for c in perfil.colunas if c.nome == "VALOR_CREDITO")
        assert col_v.null_pct >= 0.08   # ≥ 8% de nulos


# ── Par 2: Notas Fiscais de Entrada ──────────────────────────────────────────

class TestParNFe:
    def test_xlsx_row_count_4000_linhas(self, inputs_dir: Path) -> None:
        perfil = perfilar_arquivo(inputs_dir / "AUX_MOD_010_nfe_entrada.xlsx")
        assert perfil.row_count == 4000

    def test_xlsx_chave_nf_unica(self, inputs_dir: Path) -> None:
        perfil = perfilar_arquivo(inputs_dir / "AUX_MOD_010_nfe_entrada.xlsx")
        col_chave = next(c for c in perfil.colunas if c.nome == "CHAVE_NF")
        assert col_chave.alerta_duplicatas is False

    def test_xlsx_cnpj_30pct_nulos_alerta_alta(self, inputs_dir: Path) -> None:
        perfil = perfilar_arquivo(inputs_dir / "AUX_MOD_010_nfe_entrada.xlsx")
        alertas = [
            a for a in perfil.alertas
            if a.tipo == "NULOS_SIGNIFICATIVOS" and a.coluna == "CNPJ_EMITENTE"
        ]
        assert len(alertas) >= 1
        assert alertas[0].severidade == "ALTA"

    def test_csv_row_count_4000_linhas(self, inputs_dir: Path) -> None:
        perfil = perfilar_arquivo(inputs_dir / "notas_fiscais_entrada_2023.csv")
        assert perfil.row_count == 4000

    def test_csv_cnpj_30pct_nulos_alerta_alta(self, inputs_dir: Path) -> None:
        perfil = perfilar_arquivo(inputs_dir / "notas_fiscais_entrada_2023.csv")
        alertas = [
            a for a in perfil.alertas
            if a.tipo == "NULOS_SIGNIFICATIVOS" and a.coluna == "CNPJ_EMITENTE"
        ]
        assert len(alertas) >= 1
        assert alertas[0].severidade == "ALTA"


# ── Par 3: Produtor Rural PF ──────────────────────────────────────────────────

class TestParProdutorRural:
    def test_xlsx_row_count_600_linhas_resumo(self, inputs_dir: Path) -> None:
        perfil = perfilar_arquivo(inputs_dir / "AUX_MOD_010_produtor_rural.xlsx")
        assert perfil.row_count == 600

    def test_xlsx_3_abas(self, inputs_dir: Path) -> None:
        perfil = perfilar_arquivo(inputs_dir / "AUX_MOD_010_produtor_rural.xlsx")
        assert len(perfil.abas) == 3
        assert "Resumo" in perfil.abas
        assert "NCM" in perfil.abas
        assert "Parâmetros" in perfil.abas

    def test_xlsx_receita_bruta_tem_stats_numericas(self, inputs_dir: Path) -> None:
        perfil = perfilar_arquivo(inputs_dir / "AUX_MOD_010_produtor_rural.xlsx")
        col_r = next(c for c in perfil.colunas if c.nome == "RECEITA_BRUTA")
        assert col_r.tipo_inferido == "numerico"
        assert col_r.min_valor is not None
        assert col_r.max_valor is not None
        assert col_r.media_valor is not None
        # range esperado: 5.000 a 64.900
        assert float(col_r.min_valor) >= 5000.0
        assert float(col_r.max_valor) <= 65000.0

    def test_xlsx_cpf_unico(self, inputs_dir: Path) -> None:
        perfil = perfilar_arquivo(inputs_dir / "AUX_MOD_010_produtor_rural.xlsx")
        col_cpf = next(c for c in perfil.colunas if c.nome == "CPF")
        assert col_cpf.alerta_duplicatas is False

    def test_csv_row_count_600_linhas(self, inputs_dir: Path) -> None:
        perfil = perfilar_arquivo(inputs_dir / "producao_rural_pf_2023.csv")
        assert perfil.row_count == 600

    def test_csv_receita_bruta_min_max(self, inputs_dir: Path) -> None:
        perfil = perfilar_arquivo(inputs_dir / "producao_rural_pf_2023.csv")
        col_r = next(c for c in perfil.colunas if c.nome == "RECEITA_BRUTA")
        assert col_r.min_valor is not None
        assert col_r.max_valor is not None


# ── Cross-file: colunas compartilhadas ───────────────────────────────────────

class TestColunasCompartilhadas:
    def test_cpf_aparece_em_multiplos_arquivos(self, pacote) -> None:
        """CPF está em xlsx+csv do Par 1 e xlsx+csv do Par 3 → ≥ 4 arquivos."""
        cpf_entry = next(
            (c for c in pacote.colunas_compartilhadas if c["coluna"] == "CPF"), None
        )
        assert cpf_entry is not None
        assert len(cpf_entry["arquivos"]) >= 4

    def test_chave_nf_aparece_em_xlsx_e_csv(self, pacote) -> None:
        """CHAVE_NF está no xlsx e no csv do Par 2."""
        chave_entry = next(
            (c for c in pacote.colunas_compartilhadas if c["coluna"] == "CHAVE_NF"), None
        )
        assert chave_entry is not None
        assert len(chave_entry["arquivos"]) >= 2

    def test_valor_nf_compartilhado_par2(self, pacote) -> None:
        """VALOR_NF está no xlsx e no csv do Par 2."""
        v_entry = next(
            (c for c in pacote.colunas_compartilhadas if c["coluna"] == "VALOR_NF"), None
        )
        assert v_entry is not None
        assert len(v_entry["arquivos"]) >= 2

    def test_pacote_tem_alertas_criticos(self, pacote) -> None:
        """Par 1 XLSX tem CPFs duplicados → pelo menos 1 alerta crítico no pacote."""
        assert pacote.total_alertas_criticos >= 1


# ── Integração com file_prep ──────────────────────────────────────────────────

class TestIntegracaoFilePrepPerfil:
    def test_preparar_arquivo_csv_inclui_prefixo_analitico(
        self, inputs_dir: Path, perfis_dir: Path
    ) -> None:
        """
        Depois de perfilar_diretorio, preparar_arquivo com perfil_dir deve
        incluir '[PERFIL ANALÍTICO]' antes da amostra truncada.
        """
        from diogenes.agents.file_prep import preparar_arquivo

        csv_path = inputs_dir / "base_creditos_pf_2023.csv"
        conteudo = preparar_arquivo(csv_path, perfil_dir=perfis_dir)
        assert "[PERFIL ANALÍTICO" in conteudo
        assert "row_count" in conteudo
        assert "alertas" in conteudo

    def test_preparar_arquivo_xlsx_inclui_prefixo_analitico(
        self, inputs_dir: Path, perfis_dir: Path
    ) -> None:
        from diogenes.agents.file_prep import preparar_arquivo

        xlsx_path = inputs_dir / "AUX_MOD_010_creditos_pf.xlsx"
        conteudo = preparar_arquivo(xlsx_path, perfil_dir=perfis_dir)
        assert "[PERFIL ANALÍTICO" in conteudo

    def test_preparar_arquivo_sem_perfil_dir_funciona_normal(
        self, inputs_dir: Path
    ) -> None:
        """Sem perfil_dir, comportamento idêntico ao original (sem prefixo)."""
        from diogenes.agents.file_prep import preparar_arquivo

        csv_path = inputs_dir / "base_creditos_pf_2023.csv"
        conteudo = preparar_arquivo(csv_path)
        assert "[PERFIL ANALÍTICO" not in conteudo
        assert "CPF" in conteudo   # cabeçalho do CSV ainda aparece na amostra

    def test_ler_perfil_texto_retorna_yaml_valido(
        self, inputs_dir: Path, perfis_dir: Path
    ) -> None:
        import yaml

        csv_path = inputs_dir / "notas_fiscais_entrada_2023.csv"
        texto = ler_perfil_texto(perfis_dir, csv_path)
        assert texto != ""
        dados = yaml.safe_load(texto)
        assert dados["row_count"] == 4000
        assert len(dados["colunas"]) == 5
        assert len(dados["alertas"]) >= 1   # CNPJ_EMITENTE nulos
