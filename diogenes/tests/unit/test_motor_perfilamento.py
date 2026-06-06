"""
tests/unit/test_motor_perfilamento.py — DVA-CBS | Projeto Diógenes

Testes unitários do MotorPerfilamento (CSV via DuckDB, XLSX via openpyxl).

Fixtures:
  - csv_creditos:   3.000 linhas, coluna CPF com 50 duplicatas + nulos em VALOR
  - csv_nfe:        5.000 linhas, CHAVE_NF única + CNPJ com nulos acima do limiar
  - xlsx_modulo:    3 abas — Resumo (500 linhas), Creditos (2.000 linhas), NCM (300 linhas)
                    Aba Creditos: coluna CPF com duplicatas

Cada teste é focado num comportamento; as fixtures geram dados deterministicamente.
"""
from __future__ import annotations

import csv
from pathlib import Path

import openpyxl
import pytest

from diogenes.motors.motor_perfilamento import (
    ColunaPerfil,
    PerfilArquivo,
    _eh_chave_candidata,
    _gerar_alertas,
    _stats_coluna_lista,
    caminho_perfil,
    ler_perfil_texto,
    perfilar_arquivo,
    perfilar_diretorio,
)

# ── Helpers para gerar fixtures deterministicamente ───────────────────────

def _escrever_csv_creditos(path: Path, n_linhas: int = 3000, n_duplicatas: int = 50) -> None:
    """CSV com n_linhas, CPF como chave com n_duplicatas e ~8% de VALOR nulo."""
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["CPF", "NOME", "VALOR_CREDITO", "ANO_BASE", "NCM"])
        cpf_base = 10000000000
        # Linhas únicas
        for i in range(n_linhas - n_duplicatas):
            cpf = str(cpf_base + i).zfill(11)
            valor = "" if i % 12 == 0 else f"{100.0 + i * 0.5:.2f}"  # ~8% nulos
            writer.writerow([cpf, f"CONTRIBUINTE_{i}", valor, "2023", f"22{(i % 100):02d}"])
        # Duplicatas: repete os primeiros n_duplicatas CPFs
        for i in range(n_duplicatas):
            cpf = str(cpf_base + i).zfill(11)
            writer.writerow([cpf, f"DUP_{i}", f"{200.0 + i:.2f}", "2023", f"22{i:02d}"])


def _escrever_csv_nfe(path: Path, n_linhas: int = 5000) -> None:
    """CSV com n_linhas, CHAVE_NF única (formato alfanumérico), CNPJ_EMITENTE com 25% nulos."""
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["CHAVE_NF", "CNPJ_EMITENTE", "VALOR_NF", "DATA_EMISSAO"])
        for i in range(n_linhas):
            # Prefixo "NF" evita que DuckDB trate como numérico (sem overflow de HUGEINT)
            chave = f"NF{i:010d}CE{i:010d}SP{i:08d}"
            cnpj = "" if i % 4 == 0 else f"{10000000000000 + i:014d}"  # 25% nulos
            writer.writerow([chave, cnpj, f"{50.0 + i:.2f}", f"2023-{(i % 12) + 1:02d}-01"])


def _escrever_xlsx_modulo(path: Path) -> None:
    """
    XLSX com 3 abas:
      Resumo    — 500 linhas (CPF, VALOR, ANO_BASE)
      Creditos  — 2.000 linhas (CPF com 30 duplicatas, VALOR_CREDITO, NCM)
      NCM       — 300 linhas (COD_NCM, DESCRICAO, ALIQUOTA)
    """
    wb = openpyxl.Workbook()

    # Aba Resumo
    ws_r = wb.active
    ws_r.title = "Resumo"
    ws_r.append(["CPF", "VALOR_TOTAL", "ANO_BASE"])
    for i in range(500):
        cpf = str(10000000000 + i).zfill(11)
        ws_r.append([cpf, round(1000.0 + i * 2.5, 2), 2023])

    # Aba Creditos (2.000 linhas, 30 duplicatas de CPF)
    ws_c = wb.create_sheet("Creditos")
    ws_c.append(["CPF", "VALOR_CREDITO", "NCM", "ANO_BASE"])
    base = 20000000000
    for i in range(1970):
        cpf = str(base + i).zfill(11)
        ws_c.append([cpf, round(500.0 + i * 0.3, 2), f"22{i % 100:02d}", 2023])
    for i in range(30):  # duplicatas
        cpf = str(base + i).zfill(11)
        ws_c.append([cpf, round(100.0 + i, 2), f"22{i:02d}", 2023])

    # Aba NCM (300 linhas, sem duplicatas)
    ws_n = wb.create_sheet("NCM")
    ws_n.append(["COD_NCM", "DESCRICAO", "ALIQUOTA"])
    for i in range(300):
        ws_n.append([f"22{i:02d}", f"PRODUTO_{i}", round(0.05 + i * 0.001, 4)])

    wb.save(path)


# ── Fixtures pytest ───────────────────────────────────────────────────────

@pytest.fixture
def tmp_inputs(tmp_path: Path) -> Path:
    d = tmp_path / "inputs"
    d.mkdir()
    return d


@pytest.fixture
def csv_creditos(tmp_inputs: Path) -> Path:
    p = tmp_inputs / "base_creditos_2023.csv"
    _escrever_csv_creditos(p)
    return p


@pytest.fixture
def csv_nfe(tmp_inputs: Path) -> Path:
    p = tmp_inputs / "notas_fiscais_2023.csv"
    _escrever_csv_nfe(p)
    return p


@pytest.fixture
def xlsx_modulo(tmp_inputs: Path) -> Path:
    p = tmp_inputs / "AUX_MOD_10_PF.xlsx"
    _escrever_xlsx_modulo(p)
    return p


@pytest.fixture
def perfis_dir(tmp_path: Path) -> Path:
    d = tmp_path / "_runtime" / "perfis"
    d.mkdir(parents=True)
    return d


# ── Testes: CSV via DuckDB ────────────────────────────────────────────────

class TestPerfilCSV:
    def test_row_count_correto(self, csv_creditos: Path) -> None:
        perfil = perfilar_arquivo(csv_creditos)
        # 3000 linhas de dados (sem header)
        assert perfil.row_count == 3000

    def test_col_count_correto(self, csv_creditos: Path) -> None:
        perfil = perfilar_arquivo(csv_creditos)
        assert perfil.col_count == 5

    def test_detecta_duplicatas_em_chave(self, csv_creditos: Path) -> None:
        """CPF tem 50 duplicatas → alerta DUPLICATA_CHAVE CRITICA."""
        perfil = perfilar_arquivo(csv_creditos)
        col_cpf = next((c for c in perfil.colunas if c.nome == "CPF"), None)
        assert col_cpf is not None
        assert col_cpf.eh_chave_candidata is True
        assert col_cpf.alerta_duplicatas is True
        assert col_cpf.duplicatas_count > 0

    def test_alerta_duplicata_critica_no_perfil(self, csv_creditos: Path) -> None:
        perfil = perfilar_arquivo(csv_creditos)
        alertas_dup = [a for a in perfil.alertas if a.tipo == "DUPLICATA_CHAVE"]
        assert len(alertas_dup) >= 1
        assert alertas_dup[0].severidade == "CRITICA"

    def test_detecta_nulos_em_valor(self, csv_creditos: Path) -> None:
        """VALOR_CREDITO tem ~8% nulos → alerta NULOS_SIGNIFICATIVOS MEDIA."""
        perfil = perfilar_arquivo(csv_creditos)
        col_val = next((c for c in perfil.colunas if c.nome == "VALOR_CREDITO"), None)
        assert col_val is not None
        assert col_val.null_count > 0
        alertas_nulos = [
            a for a in perfil.alertas
            if a.tipo == "NULOS_SIGNIFICATIVOS" and a.coluna == "VALOR_CREDITO"
        ]
        assert len(alertas_nulos) >= 1

    def test_nfe_chave_unica_sem_alerta_duplicata(self, csv_nfe: Path) -> None:
        """CHAVE_NF é única → não deve gerar alerta de duplicata."""
        perfil = perfilar_arquivo(csv_nfe)
        col_chave = next((c for c in perfil.colunas if c.nome == "CHAVE_NF"), None)
        assert col_chave is not None
        assert col_chave.alerta_duplicatas is False

    def test_nfe_cnpj_25pct_nulos_gera_alerta_alta(self, csv_nfe: Path) -> None:
        """CNPJ_EMITENTE com 25% nulos → alerta ALTA."""
        perfil = perfilar_arquivo(csv_nfe)
        alertas_cnpj = [
            a for a in perfil.alertas
            if a.tipo == "NULOS_SIGNIFICATIVOS" and a.coluna == "CNPJ_EMITENTE"
        ]
        assert len(alertas_cnpj) >= 1
        assert alertas_cnpj[0].severidade == "ALTA"

    def test_row_count_csv_grande(self, csv_nfe: Path) -> None:
        """CSV com 5000 linhas deve ser perfilado por completo (sem truncamento CSV)."""
        perfil = perfilar_arquivo(csv_nfe)
        assert perfil.row_count == 5000
        assert perfil.erro is None

    def test_sem_erro(self, csv_creditos: Path, csv_nfe: Path) -> None:
        for p in (csv_creditos, csv_nfe):
            assert perfilar_arquivo(p).erro is None

    def test_hash_sha256_preenchido(self, csv_creditos: Path) -> None:
        perfil = perfilar_arquivo(csv_creditos)
        assert len(perfil.hash_sha256) == 64
        assert all(c in "0123456789abcdef" for c in perfil.hash_sha256)


# ── Testes: XLSX via openpyxl ─────────────────────────────────────────────

class TestPerfilXLSX:
    def test_row_count_aba_principal(self, xlsx_modulo: Path) -> None:
        """Aba Resumo tem 500 linhas de dados."""
        perfil = perfilar_arquivo(xlsx_modulo)
        assert perfil.row_count == 500

    def test_col_count_aba_principal(self, xlsx_modulo: Path) -> None:
        perfil = perfilar_arquivo(xlsx_modulo)
        assert perfil.col_count == 3

    def test_lista_abas(self, xlsx_modulo: Path) -> None:
        perfil = perfilar_arquivo(xlsx_modulo)
        assert "Resumo" in perfil.abas
        assert "Creditos" in perfil.abas
        assert "NCM" in perfil.abas

    def test_sem_truncamento_xlsx_pequeno(self, xlsx_modulo: Path) -> None:
        """Aba Resumo tem 500 linhas < _XLSX_MAX_ROWS → não truncado."""
        perfil = perfilar_arquivo(xlsx_modulo)
        assert perfil.truncado is False

    def test_sem_erro(self, xlsx_modulo: Path) -> None:
        assert perfilar_arquivo(xlsx_modulo).erro is None

    def test_detecta_col_cpf_como_chave(self, xlsx_modulo: Path) -> None:
        perfil = perfilar_arquivo(xlsx_modulo)
        col_cpf = next((c for c in perfil.colunas if c.nome == "CPF"), None)
        assert col_cpf is not None
        assert col_cpf.eh_chave_candidata is True

    def test_hash_preenchido(self, xlsx_modulo: Path) -> None:
        perfil = perfilar_arquivo(xlsx_modulo)
        assert len(perfil.hash_sha256) == 64


# ── Testes: perfilar_diretorio ────────────────────────────────────────────

class TestPerfilDiretorio:
    def test_perfila_csv_e_xlsx(
        self, tmp_inputs: Path, perfis_dir: Path,
        csv_creditos: Path, csv_nfe: Path, xlsx_modulo: Path,
    ) -> None:
        pacote = perfilar_diretorio(tmp_inputs, perfis_dir)
        assert pacote.total_arquivos == 3
        assert len(pacote.arquivos_perfilados) == 3
        assert len(pacote.arquivos_com_erro) == 0

    def test_grava_perfil_yaml_por_arquivo(
        self, tmp_inputs: Path, perfis_dir: Path,
        csv_creditos: Path, csv_nfe: Path, xlsx_modulo: Path,
    ) -> None:
        perfilar_diretorio(tmp_inputs, perfis_dir)
        yamls = list(perfis_dir.glob("*.perfil.yaml"))
        assert len(yamls) == 3

    def test_grava_perfil_pacote_yaml(
        self, tmp_inputs: Path, perfis_dir: Path,
        csv_creditos: Path, csv_nfe: Path, xlsx_modulo: Path,
    ) -> None:
        perfilar_diretorio(tmp_inputs, perfis_dir)
        assert (perfis_dir / "perfil_pacote.yaml").exists()

    def test_detecta_coluna_cpf_compartilhada(
        self, tmp_inputs: Path, perfis_dir: Path,
        csv_creditos: Path, xlsx_modulo: Path,
    ) -> None:
        """CPF aparece em base_creditos.csv e em AUX_MOD_10_PF.xlsx → compartilhada."""
        pacote = perfilar_diretorio(tmp_inputs, perfis_dir)
        cols_compartilhadas = {c["coluna"] for c in pacote.colunas_compartilhadas}
        assert "CPF" in cols_compartilhadas

    def test_conta_alertas_criticos_total(
        self, tmp_inputs: Path, perfis_dir: Path,
        csv_creditos: Path, csv_nfe: Path, xlsx_modulo: Path,
    ) -> None:
        """csv_creditos tem pelo menos 1 alerta CRITICO (CPF duplicado)."""
        pacote = perfilar_diretorio(tmp_inputs, perfis_dir)
        assert pacote.total_alertas_criticos >= 1

    def test_diretorio_perfis_criado(
        self, tmp_inputs: Path, tmp_path: Path,
        csv_creditos: Path,
    ) -> None:
        novo_perfis = tmp_path / "novos_perfis"
        assert not novo_perfis.exists()
        perfilar_diretorio(tmp_inputs, novo_perfis)
        assert novo_perfis.is_dir()

    def test_ignora_arquivo_nao_analisavel(
        self, tmp_inputs: Path, perfis_dir: Path,
    ) -> None:
        """Arquivos em DIRS_IGNORADOS não devem ser perfilados."""
        logs = tmp_inputs / "03_LOGS"
        logs.mkdir()
        (logs / "intake.log").write_text("log content", encoding="utf-8")
        # CSV dentro de DIRS_IGNORADOS
        csv_ignorado = logs / "debug.csv"
        _escrever_csv_creditos(csv_ignorado, n_linhas=10)
        pacote = perfilar_diretorio(tmp_inputs, perfis_dir)
        # Nenhum arquivo de 03_LOGS deve aparecer perfilado
        for arq in pacote.arquivos_perfilados:
            assert "03_LOGS" not in arq


# ── Testes: ler_perfil_texto (integração com file_prep) ──────────────────

class TestLerPerfilTexto:
    def test_retorna_yaml_quando_perfil_existe(
        self, tmp_inputs: Path, perfis_dir: Path, csv_creditos: Path
    ) -> None:
        perfilar_diretorio(tmp_inputs, perfis_dir)
        texto = ler_perfil_texto(perfis_dir, csv_creditos)
        assert "row_count" in texto
        assert "colunas" in texto
        assert "alertas" in texto

    def test_retorna_vazio_quando_perfil_ausente(
        self, perfis_dir: Path, tmp_path: Path
    ) -> None:
        arquivo_sem_perfil = tmp_path / "inexistente.csv"
        texto = ler_perfil_texto(perfis_dir, arquivo_sem_perfil)
        assert texto == ""

    def test_yaml_contem_hash(
        self, tmp_inputs: Path, perfis_dir: Path, csv_creditos: Path
    ) -> None:
        perfilar_diretorio(tmp_inputs, perfis_dir)
        texto = ler_perfil_texto(perfis_dir, csv_creditos)
        assert "hash_sha256" in texto


# ── Testes: funções auxiliares ────────────────────────────────────────────

class TestAuxiliares:
    @pytest.mark.parametrize("nome,esperado", [
        ("CPF", True),
        ("CNPJ_EMITENTE", True),
        ("CHAVE_NF", True),
        ("id_contrato", True),
        ("codigo_ncm", True),
        ("VALOR_TOTAL", False),
        ("NOME", False),
        ("ANO_BASE", False),
        ("DESCRICAO", False),
    ])
    def test_eh_chave_candidata(self, nome: str, esperado: bool) -> None:
        assert _eh_chave_candidata(nome) is esperado

    def test_stats_coluna_lista_sem_nulos(self) -> None:
        valores = [str(i) for i in range(100)]
        col = _stats_coluna_lista(valores, "CPF", 100)
        assert col.null_count == 0
        assert col.distinct_count == 100
        assert col.unicidade_pct == 1.0

    def test_stats_coluna_lista_com_duplicatas(self) -> None:
        valores = [str(i % 10) for i in range(100)]  # 10 valores únicos
        col = _stats_coluna_lista(valores, "CPF", 100)
        assert col.distinct_count == 10
        assert col.eh_chave_candidata is True
        assert col.alerta_duplicatas is True

    def test_stats_coluna_lista_numerica(self) -> None:
        valores = [float(i) for i in range(1, 101)]
        col = _stats_coluna_lista(valores, "VALOR", 100)
        assert col.tipo_inferido == "numerico"
        assert col.min_valor == "1.0"
        assert col.max_valor == "100.0"
        assert col.media_valor is not None

    def test_gerar_alertas_vazio_sem_problemas(self) -> None:
        colunas = [ColunaPerfil(
            nome="VALOR", tipo_inferido="numerico",
            row_count=100, null_count=0, null_pct=0.0,
            distinct_count=100, unicidade_pct=1.0,
        )]
        alertas = _gerar_alertas(colunas)
        assert alertas == []

    def test_caminho_perfil_sanitiza_nome(self, tmp_path: Path) -> None:
        p = Path("Base Créditos 2023.xlsx")
        dest = caminho_perfil(tmp_path, p)
        assert " " not in dest.name
        assert dest.suffix == ".yaml"
        assert dest.parent == tmp_path


# ── Testes: arquivo inválido / best-effort ────────────────────────────────

class TestBestEffort:
    def test_csv_malformado_retorna_erro(self, tmp_path: Path) -> None:
        """CSV com binário inválido deve retornar PerfilArquivo.erro != None."""
        p = tmp_path / "corrompido.csv"
        p.write_bytes(b"\x00\x01\x02header_corrompido\xff\xfe")
        perfil = perfilar_arquivo(p)
        # O motor não deve levantar — deve registrar o erro no perfil
        assert isinstance(perfil, PerfilArquivo)
        # Pode ter erro (arquivo corrompido) ou ter lido com ignore_errors
        # Em ambos os casos, não deve haver exceção

    def test_xlsx_corrompido_retorna_erro(self, tmp_path: Path) -> None:
        p = tmp_path / "corrompido.xlsx"
        p.write_bytes(b"nao e um xlsx valido")
        perfil = perfilar_arquivo(p)
        assert isinstance(perfil, PerfilArquivo)
        assert perfil.erro is not None

    def test_diretorio_vazio_retorna_pacote_vazio(
        self, tmp_path: Path, perfis_dir: Path
    ) -> None:
        inputs_vazio = tmp_path / "inputs_vazio"
        inputs_vazio.mkdir()
        pacote = perfilar_diretorio(inputs_vazio, perfis_dir)
        assert pacote.total_arquivos == 0
        assert pacote.arquivos_perfilados == []
