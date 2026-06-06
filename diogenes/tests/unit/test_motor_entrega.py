"""
tests/unit/test_motor_entrega.py — Fase de Entrega

Cobre o caminho determinístico (sem LLM e sem dependências pesadas obrigatórias):
parsing do JSON do Sherlock, extração financeira via openpyxl, render do dashboard,
mapeamento do apêndice, montagem do PacoteEntrega e o Motor de Entrega ponta a ponta.
"""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pytest

from diogenes.delivery import builders, dashboard, parsing
from diogenes.delivery.extractor import ExtractorFinanceiro
from diogenes.delivery.pacote import montar_pacote
from diogenes.models import (
    BlocoFinanceiro,
    CenarioSensibilidade,
    GraficoEntrega,
    KPIEntrega,
    MetodologiaCard,
    PacoteEntrega,
    TabelaEntrega,
)
from diogenes.models import TesteRealizado as _Teste  # alias: evita coleta pelo pytest

_SHERLOCK_MD = """
## 11. JSON de Ocorrências para o Dashboard
```json
{"modulo":"MOD_010","ciclo":"C1","ocorrencias":[
 {"codigo":"S001-DIV","titulo":"Divergência de escopo","nivel":"CRITICO",
  "fundamento_violado":"LC 214/2025 art. 5","descricao":"texto","solicitacao_rfb":"demonstrar X","status":"aberto"},
 {"codigo":"S002-AT","titulo":"Ponto de atenção","nivel":"ATENCAO","descricao":"d2","status":"encaminhado"}],
 "pendencias_simulador":[{"codigo":"PC-01","descricao":"integrar simulador","verificacao_futura":"recalculo"}]}
```
### 10.4 Aderência metodológica
Texto da aderência.
### 10.5 Posição Consolidada
APROVADO_COM_RESSALVAS
"""


# ── parsing ─────────────────────────────────────────────────────

def test_extrair_json_dashboard_e_ocorrencias():
    obj = parsing.extrair_json_dashboard(_SHERLOCK_MD)
    assert obj is not None
    ocs, pend = parsing.ocorrencias_de_json(obj)
    assert [o.codigo for o in ocs] == ["S001-DIV", "S002-AT"]
    assert ocs[0].nivel == "CRITICO" and ocs[1].nivel == "ATENCAO"
    assert pend[0].codigo == "PC-01"


def test_extrair_json_tolera_virgula_final():
    obj = parsing.extrair_json_dashboard('```json\n{"ocorrencias":[],}\n```')
    assert obj == {"ocorrencias": []}


def test_extrair_veredito_e_secao():
    assert parsing.extrair_veredito(_SHERLOCK_MD) == "APROVADO_COM_RESSALVAS"
    assert "aderência" in parsing.extrair_secao(_SHERLOCK_MD, "10.4").lower()


def test_nivel_invalido_vira_atencao():
    ocs, _ = parsing.ocorrencias_de_json(
        {"ocorrencias": [{"codigo": "X", "titulo": "t", "nivel": "ZZZ"}]})
    assert ocs[0].nivel == "ATENCAO"


# ── extrator financeiro (números determinísticos) ───────────────

@pytest.fixture
def planilha(tmp_path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumo"
    ws["B3"] = 30_840_000_000          # arrecadação
    ws["B12"] = 20_060_000_000         # débitos 2023
    ws["C12"] = 19_890_000_000         # débitos 2024
    sens = wb.create_sheet("Sensibilidade")
    for i, (r, v) in enumerate(zip([0, 10, 20, 30, 40], [47.86, 43.08, 38.29, 33.5, 28.72], strict=True), start=2):
        sens[f"A{i}"] = r
        sens[f"B{i}"] = v
    path = tmp_path / "AUX_MOD_10.xlsx"
    wb.save(path)
    return path


def test_extrator_le_valor_exato_e_formata(planilha: Path):
    mapa = {
        "blocos": [{
            "id": "vg", "titulo": "Visão Geral",
            "kpis": [{"rotulo": "Arrecadação", "aba": "Resumo", "celula": "B3", "formato": "bilhoes"}],
        }],
        "valores_agregados": [
            {"descricao": "Débitos", "aba": "Resumo", "celula_2023": "B12",
             "celula_2024": "C12", "formato": "bilhoes"},
        ],
        "sensibilidade_redutor": {"aba": "Sensibilidade",
                                  "intervalo_redutores": "A2:A6", "intervalo_valores": "B2:B6"},
    }
    res = ExtractorFinanceiro(planilha).extrair(mapa)
    assert res["blocos_financeiros"][0].kpis[0].valor == "R$ 30,84 bilhões"
    assert res["valores_agregados"][0]["valor_2023"] == "R$ 20,06 bilhões"
    assert res["sensibilidade_redutor"]["redutores"] == [0, 10, 20, 30, 40]
    assert res["sensibilidade_redutor"]["valores"][0] == 47.86


def test_extrator_aba_ausente_gera_aviso(planilha: Path):
    ext = ExtractorFinanceiro(planilha)
    res = ext.extrair({"blocos": [{"id": "x", "titulo": "X",
                                   "kpis": [{"rotulo": "k", "aba": "Inexistente", "celula": "A1"}]}]})
    assert res["blocos_financeiros"][0].kpis[0].valor == ""
    assert any("Inexistente" in a for a in ext.avisos)


def test_extrator_le_delta_total_labels_e_cenarios(planilha: Path):
    mapa = {"blocos": [
        {"id": "visao_geral", "tipo": "visao_geral", "titulo": "VG",
         "kpis": [{"rotulo": "Débitos", "aba": "Resumo", "celula": "C12",
                   "celula_base": "B12", "formato": "bilhoes", "unidade": "R$ bi"}]},
        {"id": "an", "tipo": "analitica", "titulo": "An",
         "tabelas": [{"titulo": "T", "aba": "Resumo", "intervalo": "B3:C3",
                      "cabecalho_na_primeira_linha": False, "total_labels": ["Total"]}]},
        {"id": "sens", "tipo": "sensibilidade", "titulo": "S",
         "cenarios": [{"rotulo": "Base", "aba": "Sensibilidade", "celula": "B2",
                       "formato": "texto", "destaque": "navy"}]},
    ]}
    res = ExtractorFinanceiro(planilha).extrair(mapa)
    blocos = res["blocos_financeiros"]
    kpi = blocos[0].kpis[0]
    assert kpi.unidade == "R$ bi"
    assert kpi.delta == "-0,8%" and kpi.delta_tipo == "down"
    assert blocos[1].tipo == "analitica"
    assert blocos[1].tabelas[0].total_labels == ["Total"]
    assert blocos[2].tipo == "sensibilidade"
    assert blocos[2].cenarios[0].rotulo == "Base" and blocos[2].cenarios[0].valor == "47,86"


# ── dashboard v5: visão geral, abas analíticas, sensibilidade ────

def _pacote_v5() -> PacoteEntrega:
    ocs, pend = parsing.ocorrencias_de_json(parsing.extrair_json_dashboard(_SHERLOCK_MD))
    vg = BlocoFinanceiro(
        id="visao_geral", titulo="Visão Geral", tipo="visao_geral",
        narrativa="Síntese do módulo.",
        kpis=[KPIEntrega("Arrecadação líquida", "R$ 30,84 bilhões", unidade="R$ bi",
                         delta="+3,1%", delta_tipo="up", destaque="green")],
        cards_metodologia=[MetodologiaCard(
            tag="Seção 10.1", titulo="Produtor Rural",
            corpo="Faturamento > R$ 3,6 mi via DIRPF.", chips=["DIRPF", "Arts. 164/165"])],
        graficos=[GraficoEntrega("barras", "Resumo", ["A", "B"], [1.0, 2.0])])
    an = BlocoFinanceiro(
        id="pr", titulo="Produtor Rural 10.1", tipo="analitica", narrativa="Decomposição.",
        tabelas=[TabelaEntrega("Composição", ["Item", "Tipo", "2024"],
                               [["Receita Bruta", "Débito", "547,30"],
                                ["Arrecadação líquida", "Total", "35,40"]],
                               total_labels=["Arrecadação líquida"])])
    se = BlocoFinanceiro(
        id="sens", titulo="Sensibilidade 20%", tipo="sensibilidade",
        cenarios=[CenarioSensibilidade("Homologado", "R$ 35,4 bi", destaque="navy"),
                  CenarioSensibilidade("Redutor 20%", "R$ 28,3 bi", destaque="amber")])
    return PacoteEntrega(
        cycle_id="MOD_010_A1_X", modulo=10, modulo_nome="Pessoa Física", versao="1.0",
        veredito="APROVADO_COM_RESSALVAS", blocos_financeiros=[vg, an, se],
        ocorrencias=ocs, pendencias_simulador=pend)


def test_dashboard_v5_visao_geral_metod_card_e_kpi_delta():
    h = dashboard.gerar_dashboard_html(_pacote_v5())
    assert "metod-card" in h and "Produtor Rural" in h and "DIRPF" in h
    assert "kpi-delta d-up" in h and "+3,1%" in h


def test_dashboard_v5_tabela_badge_e_total_row():
    h = dashboard.gerar_dashboard_html(_pacote_v5())
    assert 'class="badge b-blue"' in h          # Débito → badge azul
    assert 'class="total-row"' in h             # linha de Arrecadação líquida


def test_dashboard_v5_sensibilidade_scenario_cards():
    h = dashboard.gerar_dashboard_html(_pacote_v5())
    assert "scn-grid" in h and "Redutor 20%" in h and "Homologado" in h


# ── dashboard + builders ────────────────────────────────────────

def _pacote_exemplo() -> PacoteEntrega:
    ocs, pend = parsing.ocorrencias_de_json(parsing.extrair_json_dashboard(_SHERLOCK_MD))
    return PacoteEntrega(
        cycle_id="MOD_010_A1_X", modulo=10, modulo_nome="Pessoa Física", versao="1.0",
        veredito="APROVADO_COM_RESSALVAS",
        valores_agregados=[{"descricao": "Débitos", "valor_2023": "R$ 20,06 bi", "valor_2024": "R$ 19,89 bi"}],
        sensibilidade_redutor={"redutores": [0, 10], "valores": [47.86, 43.08]},
        blocos_financeiros=[BlocoFinanceiro(
            id="vg", titulo="Produtor Rural", kpis=[KPIEntrega("Arrecadação", "R$ 30,84 bi")],
            tabelas=[TabelaEntrega("Comp", ["A", "B"], [["1", "2"]], "DIRPF")])],
        ocorrencias=ocs, pendencias_simulador=pend,
    )


def test_dashboard_contem_inconsistencias_e_grafico():
    h = dashboard.gerar_dashboard_html(_pacote_exemplo())
    assert "Inconsistências" in h
    assert "S001-DIV" in h and "data-chart" in h
    assert "APROVADO_COM_RESSALVAS" in h


def test_dados_apendice_mapeia_status_e_conclusao():
    ap = builders.dados_apendice(_pacote_exemplo())
    incs = {i["id"]: i["status"] for i in ap["inconsistencias"]}
    assert incs["S001-DIV"] == "Divergência"
    assert incs["S002-AT"] == "Atendido Parcialmente"
    assert "sensibilidade_redutor" in ap["conclusao"]


def test_ficha_html_tem_duas_paginas():
    h = builders.gerar_ficha_html(_pacote_exemplo())
    assert h.count('class="page"') == 2
    assert "page-break-after" in h


# ── Apêndice: conteúdo redigido (Mycroft) + números determinísticos ─

def _conteudo_apendice() -> dict:
    return {
        "proposta": {"descricao": "Apura CBS de PF.", "fonte": "Anexo Nota Cetad 079/2025 (peça 16)"},
        "objetivo": "Estimar arrecadação própria e ajuste no MC.",
        "testes": {
            "camada_1": {
                "conformidade": [{"id": "C1-01", "descricao": "Aderência", "resultado": "Confere", "status": "Atendido"}],
                "sensibilidade": [{"id": "C1-30", "descricao": "Redutor 20%", "resultado": "Coerente", "status": "Atendido"}],
            },
            "camada_2": {"premissas": [{"id": "C2-10", "descricao": "Percentual presumido", "resultado": "Fundamentado", "status": "Atendido Parcialmente"}]},
            "camada_3": {"recalculo": [{"id": "C3-10", "descricao": "Recálculo", "resultado": "Confere", "status": "Atendido"}]},
        },
        "inconsistencias": [
            {"id": "S001-DIV", "consequencia": "Superestima base na cadeia",
             "tratamento": "RFB demonstrará o ajuste na próxima entrega"},
        ],
        "alteracoes_metodologicas": [{"id": "ALT-01", "descricao": "Inclusão do redutor",
                                      "acordo": "Acordado", "impacto": "Reduz arrecadação"}],
        "conclusao": {"conformidade": "Conforme com ressalvas.",
                      "premissas": [{"nome": "Redutor 20%", "descricao": "Alteração de comportamento",
                                     "impacto": "Reduz a base"}]},
    }


def _pacote_apendice(conteudo: dict | None) -> PacoteEntrega:
    ocs, _ = parsing.ocorrencias_de_json(parsing.extrair_json_dashboard(_SHERLOCK_MD))
    return PacoteEntrega(
        cycle_id="C", modulo=10, modulo_nome="Pessoa Física", veredito="APROVADO_COM_RESSALVAS",
        ocorrencias=ocs,
        valores_agregados=[{"descricao": "Débitos", "valor_2023": "R$ 20,06 bilhões",
                            "valor_2024": "R$ 19,89 bilhões"}],
        testes_camada_1=[_Teste("T-FB", "fallback", "ok", "Atendido")],
        apendice_conteudo=conteudo,
    )


def test_dados_apendice_mescla_conteudo_redigido():
    d = builders.dados_apendice(_pacote_apendice(_conteudo_apendice()))
    assert d["proposta"]["fonte"].startswith("Anexo Nota Cetad")
    assert set(d["testes"]["camada_1"]) >= {"conformidade", "sensibilidade"}
    assert len(d["testes"]["camada_2"]["premissas"]) == 1
    assert len(d["testes"]["camada_3"]["recalculo"]) == 1
    inc = {i["id"]: i for i in d["inconsistencias"]}
    assert inc["S001-DIV"]["consequencia"] == "Superestima base na cadeia"
    assert inc["S001-DIV"]["status"] == "Divergência"          # determinístico (nível CRITICO)
    assert len(d["alteracoes_metodologicas"]) == 1
    assert len(d["conclusao"]["premissas"]) == 1


def test_dados_apendice_numeros_sao_deterministicos():
    # Conteúdo redigido NÃO carrega números — valores agregados vêm do PacoteEntrega.
    d = builders.dados_apendice(_pacote_apendice(_conteudo_apendice()))
    assert d["conclusao"]["valores_agregados"][0]["valor_2023"] == "R$ 20,06 bilhões"


def test_dados_apendice_fallback_sem_conteudo():
    d = builders.dados_apendice(_pacote_apendice(None))
    assert list(d["testes"]["camada_1"]) == ["conformidade"]
    assert d["testes"]["camada_1"]["conformidade"][0]["id"] == "T-FB"
    assert d["alteracoes_metodologicas"] == []
    assert d["conclusao"]["premissas"] == []


def test_apendice_docx_gerado_ponta_a_ponta(tmp_path: Path):
    pytest.importorskip("docx")
    out = builders.gerar_apendice_docx(_pacote_apendice(_conteudo_apendice()),
                                       tmp_path / "Apendice.docx")
    assert out.exists() and out.stat().st_size > 0


def test_montar_pacote_carrega_apendice_conteudo(tmp_path: Path):
    cid = "MOD_010_A1_20260604T140000Z"
    cdir = tmp_path / "cycles" / cid
    (cdir / "output").mkdir(parents=True)
    (cdir / "inputs").mkdir(parents=True)
    (cdir / "output" / f"relatorio_higienizado_{cid}.md").write_text(_SHERLOCK_MD, encoding="utf-8")
    (cdir / "output" / "entrega_apendice.json").write_text(
        json.dumps(_conteudo_apendice()), encoding="utf-8")
    pacote, _ = montar_pacote(cdir, "MOD_010", 1)
    assert pacote.apendice_conteudo is not None
    assert pacote.apendice_conteudo["proposta"]["fonte"].startswith("Anexo Nota Cetad")


# ── montagem do PacoteEntrega a partir do diretório do ciclo ─────

def test_montar_pacote_le_json_e_extrai_financeiro(tmp_path: Path, planilha: Path):
    cid = "MOD_010_A1_20260604T120000Z"
    cdir = tmp_path / "cycles" / cid
    (cdir / "output").mkdir(parents=True)
    (cdir / "inputs").mkdir(parents=True)
    (cdir / "output" / f"relatorio_higienizado_{cid}.md").write_text(_SHERLOCK_MD, encoding="utf-8")
    obj = parsing.extrair_json_dashboard(_SHERLOCK_MD)
    (cdir / "output" / "sherlock_ocorrencias.json").write_text(
        json.dumps(obj), encoding="utf-8")
    # copiar planilha para inputs
    (cdir / "inputs" / "AUX_MOD_10.xlsx").write_bytes(planilha.read_bytes())
    mapa = {
        "modulo_nome": "Pessoa Física", "planilha_principal": "AUX_MOD_10.xlsx",
        "narrativa": {"objetivo": "Calcular CBS PF"},
        "valores_agregados": [{"descricao": "Débitos", "aba": "Resumo",
                               "celula_2023": "B12", "celula_2024": "C12", "formato": "bilhoes"}],
    }
    (cdir / "output" / "entrega_mapa_extracao.json").write_text(
        json.dumps(mapa), encoding="utf-8")

    pacote, avisos = montar_pacote(cdir, "MOD_010", 1)
    assert pacote.modulo == 10 and pacote.modulo_nome == "Pessoa Física"
    assert pacote.objetivo == "Calcular CBS PF"
    assert pacote.veredito == "APROVADO_COM_RESSALVAS"
    assert len(pacote.ocorrencias) == 2
    assert pacote.valores_agregados[0]["valor_2023"] == "R$ 20,06 bilhões"


# ── Motor de Entrega ponta a ponta (degrada sem deps pesadas) ────

def test_motor_entrega_gera_dashboard_e_manifesto(tmp_path, monkeypatch):
    from diogenes.config import get_config
    from diogenes.models import CycleRecord
    from diogenes.motors.motor_entrega import MotorEntrega
    from diogenes.persistence.audit_index import AuditIndex

    ws = tmp_path / "workspace"
    (ws / "cycles").mkdir(parents=True)
    monkeypatch.setenv("DIOGENES_WORKSPACE", str(ws))
    monkeypatch.setenv("DIOGENES_LLM_PROVIDER", "openrouter")
    monkeypatch.setenv("DIOGENES_LLM_API_KEY", "k")
    get_config.cache_clear()

    cid = "MOD_010_A1_20260604T130000Z"
    cdir = ws / "cycles" / cid
    (cdir / "output").mkdir(parents=True)
    (cdir / "inputs").mkdir(parents=True)
    (cdir / "output" / f"relatorio_higienizado_{cid}.md").write_text(_SHERLOCK_MD, encoding="utf-8")
    (cdir / "output" / "sherlock_ocorrencias.json").write_text(
        json.dumps(parsing.extrair_json_dashboard(_SHERLOCK_MD)), encoding="utf-8")

    audit = AuditIndex(ws)
    audit.create_if_not_exists()
    audit.add_cycle(CycleRecord(cycle_id=cid, module_id="MOD_010", activity=1,
                                status="AGUARDANDO_CHANCELA_LESTRADE", opened_at_utc="x"))

    report = MotorEntrega().gerar(cid, with_assets=False)

    by_type = {a.tipo: a for a in report.artefatos}
    assert by_type["dashboard"].ok is True
    assert by_type["ficha"].ok is True
    assert (report.entrega_dir / "Dashboard.html").exists()
    assert (report.entrega_dir / "entrega_manifesto.json").exists()
    manifesto = json.loads((report.entrega_dir / "entrega_manifesto.json").read_text(encoding="utf-8"))
    assert manifesto["contagem"]["ocorrencias"] == 2
    # audit registrou a entrega
    assert audit.get_cycle(cid)["entrega_invocado_at_utc"] != ""
