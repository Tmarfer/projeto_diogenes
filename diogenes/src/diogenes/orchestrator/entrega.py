"""
orchestrator/entrega.py — DVA-CBS | Projeto Diógenes

Orquestração da Fase de Entrega (excursão a partir de AGUARDANDO_CHANCELA_LESTRADE):

    Mycroft.mapear_dados_modulo()  → output/entrega_mapa_extracao.json   (LLM, localiza dados)
    MotorEntrega.gerar()           → output/entrega/*                    (determinístico)
    Mycroft.avaliar_entrega()      → veredito QA (APROVADO | REQUER_AJUSTE)

A geração determinística (Motor de Entrega) sempre roda. As etapas com LLM
(mapeamento + QA) degradam graciosamente: se o cliente LLM não puder ser
instanciado (ex.: ChatTCU sem autenticação) ou `rodar_qa=False`, a entrega
prossegue só com a parte determinística.
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from diogenes.config import get_config
from diogenes.delivery.pacote import NOME_APENDICE_CONTEUDO, NOME_MAPA_EXTRACAO
from diogenes.models import MotorEntregaReport
from diogenes.motors.motor_entrega import MotorEntrega
from diogenes.orchestrator.states import TRANSICOES_VALIDAS, CycleState
from diogenes.persistence.audit_index import AuditIndex
from diogenes.persistence.workspace import WorkspaceManager

logger = logging.getLogger(__name__)

_MAX_PREVIEW_LINHAS = 8
_MAX_ABAS = 30


def executar_entrega(
    cycle_id: str, *, rodar_qa: bool = True, with_assets: bool = True,
) -> MotorEntregaReport:
    cfg = get_config()
    ws = Path(cfg.workspace.path)
    audit = AuditIndex(ws)
    wm = WorkspaceManager(ws)
    cycle_dir = wm.get_cycle_dir(cycle_id)
    record = audit.get_cycle(cycle_id)
    if record is None:
        raise KeyError(f"Ciclo '{cycle_id}' não encontrado no audit_index.")

    _transicao_segura(audit, cycle_id, CycleState.EM_EXECUCAO_ENTREGA)

    mycroft = _construir_mycroft(cycle_id, cycle_dir) if rodar_qa else None

    # 1. Mapeamento de dados (só se ainda não houver mapa e o LLM estiver disponível)
    mapa_path = cycle_dir / "output" / NOME_MAPA_EXTRACAO
    if mycroft is not None and not mapa_path.exists():
        _mapear(mycroft, cycle_dir, record, mapa_path)

    # 1b. Redação do conteúdo do apêndice (reorganiza o consolidado validado)
    apendice_path = cycle_dir / "output" / NOME_APENDICE_CONTEUDO
    if mycroft is not None and not apendice_path.exists():
        _redigir_apendice(mycroft, cycle_dir, record, apendice_path)

    # 2. Geração determinística
    report = MotorEntrega().gerar(cycle_id, with_assets=with_assets)

    # 3. QA de Mycroft
    veredito = ""
    if mycroft is not None:
        veredito = _avaliar(mycroft, report)
        try:
            audit.update_entrega(cycle_id, report.gerado_em_utc,
                                 report.total_artefatos, veredito)
        except Exception as exc:  # noqa: BLE001
            logger.warning("audit.update_entrega (veredito) falhou: %s", exc)

    # 4. Retorno de estado
    if veredito == "REQUER_AJUSTE":
        _transicao_segura(audit, cycle_id, CycleState.AGUARDANDO_AJUSTE_ENTREGA)
    else:
        _transicao_segura(audit, cycle_id, CycleState.AGUARDANDO_CHANCELA_LESTRADE)

    try:
        from diogenes.reports.render_html import atualizar_report_html
        atualizar_report_html(cycle_id, ws)
    except Exception:
        pass

    return report


# ── etapas com LLM ──────────────────────────────────────────────

def _construir_mycroft(cycle_id: str, cycle_dir: Path) -> Any:
    try:
        from diogenes.agents.mycroft import MycrooftAgent
        from diogenes.llm.base import get_llm_client
        cfg = get_config()
        llm = get_llm_client(cycle_id, cycle_dir / "_runtime")
        return MycrooftAgent(
            llm=llm, agent_spec=cfg.agentes.mycroft, cycle_id=cycle_id,
            docs_dir=Path("docs/agentes") / "mycroft", cycle_dir=cycle_dir,
            seed_base=cfg.agentes.seed_base,
        )
    except Exception as exc:  # noqa: BLE001
        logger.warning("LLM indisponível para a Fase de Entrega — só geração determinística: %s", exc)
        return None


def _mapear(mycroft: Any, cycle_dir: Path, record: dict[str, Any], mapa_path: Path) -> None:
    import json

    from diogenes.models import CycleManifest
    inventario = _inventario_planilha(cycle_dir)
    if not inventario:
        logger.warning("Sem planilha em inputs/ — pulando mapeamento de dados.")
        return
    metodologia = _ler_metodologia(cycle_dir)
    ocorrencias_resumo = _resumo_ocorrencias(cycle_dir)
    # Manifesto mínimo para a chamada (mapear só usa module_id/activity)
    manifest = CycleManifest(
        cycle_id=record["cycle_id"], module_id=record.get("module_id", ""),
        activity=int(record.get("activity", "1") or "1"),
        opened_at_utc=record.get("opened_at_utc", ""), is_sigilo_module=False,
        input_files=[], package_hash="", git_commit="", diogenes_version="",
        python_version="", openai_version="", cycle_num=0,
    )
    mapa = mycroft.mapear_dados_modulo(
        manifest, inventario, metodologia=metodologia, ocorrencias_resumo=ocorrencias_resumo,
    )
    if mapa:
        mapa_path.parent.mkdir(parents=True, exist_ok=True)
        mapa_path.write_text(json.dumps(mapa, ensure_ascii=False, indent=2), encoding="utf-8")
        logger.info("Mapa de extração gravado em %s", mapa_path)


def _redigir_apendice(mycroft: Any, cycle_dir: Path, record: dict[str, Any], apendice_path: Path) -> None:
    import json

    from diogenes.models import CycleManifest
    consolidado = _ler_consolidado(cycle_dir)
    if not consolidado:
        logger.warning("Relatório consolidado ausente — pulando redação do apêndice.")
        return
    manifest = CycleManifest(
        cycle_id=record["cycle_id"], module_id=record.get("module_id", ""),
        activity=int(record.get("activity", "1") or "1"),
        opened_at_utc=record.get("opened_at_utc", ""), is_sigilo_module=False,
        input_files=[], package_hash="", git_commit="", diogenes_version="",
        python_version="", openai_version="", cycle_num=0,
    )
    conteudo = mycroft.redigir_apendice(
        manifest, consolidado,
        ocorrencias_resumo=_resumo_ocorrencias(cycle_dir),
        metodologia=_ler_metodologia(cycle_dir),
    )
    if conteudo:
        apendice_path.parent.mkdir(parents=True, exist_ok=True)
        apendice_path.write_text(json.dumps(conteudo, ensure_ascii=False, indent=2), encoding="utf-8")
        logger.info("Conteúdo do apêndice gravado em %s", apendice_path)


def _ler_consolidado(cycle_dir: Path) -> str:
    output_dir = cycle_dir / "output"
    for padrao in ("relatorio_higienizado_*.md", "relatorio_*.md"):
        achados = sorted(output_dir.glob(padrao))
        if achados:
            return achados[0].read_text(encoding="utf-8", errors="replace")
    return ""


def _avaliar(mycroft: Any, report: MotorEntregaReport) -> str:
    import contextlib
    import json
    manifesto_path = report.entrega_dir / "entrega_manifesto.json"
    manifesto: dict[str, Any] = {}
    if manifesto_path.exists():
        with contextlib.suppress(json.JSONDecodeError):
            manifesto = json.loads(manifesto_path.read_text(encoding="utf-8"))
    avaliacao = mycroft.avaliar_entrega(manifesto)
    return "APROVADO" if avaliacao.tipo == "APROVADO" else "REQUER_AJUSTE"


def _inventario_planilha(cycle_dir: Path) -> str:
    """Monta um inventário textual das abas da planilha principal (preview de cada aba)."""
    inputs = cycle_dir / "inputs"
    xlsx = sorted(inputs.rglob("*.xlsx")) if inputs.is_dir() else []
    if not xlsx:
        return ""
    path = xlsx[0]
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        logger.warning("Falha ao abrir planilha para inventário: %s", exc)
        return f"Planilha: {path.name} (não foi possível ler as abas)"
    linhas = [f"Planilha principal: {path.name}", ""]
    for nome in wb.sheetnames[:_MAX_ABAS]:
        ws = wb[nome]
        linhas.append(f"### Aba '{nome}' (dim {ws.calculate_dimension() if not ws.max_row else f'{ws.max_row}x{ws.max_column}'})")
        for i, row in enumerate(ws.iter_rows(max_row=_MAX_PREVIEW_LINHAS, values_only=True)):
            cells = " | ".join("" if c is None else str(c)[:24] for c in row[:8])
            linhas.append(f"  L{i+1}: {cells}")
        linhas.append("")
    return "\n".join(linhas)


_MAX_METODOLOGIA_CHARS = 24000


def _ler_metodologia(cycle_dir: Path) -> str:
    """Texto da metodologia homologada do módulo, para Mycroft curar os cards/narrativas."""
    inputs = cycle_dir / "inputs"
    if not inputs.is_dir():
        return ""
    candidatos = sorted(inputs.rglob("[Mm]etodologia*.md")) or sorted(inputs.rglob("*[Mm]etodolog*.md"))
    if not candidatos:
        return ""
    texto = candidatos[0].read_text(encoding="utf-8", errors="replace")
    return texto[:_MAX_METODOLOGIA_CHARS]


def _resumo_ocorrencias(cycle_dir: Path) -> str:
    """Resumo textual das ocorrências do Sherlock (códigos/níveis/títulos), sem dados brutos."""
    import json

    jpath = cycle_dir / "output" / "sherlock_ocorrencias.json"
    if not jpath.exists():
        return ""
    try:
        obj = json.loads(jpath.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return ""
    itens = obj.get("ocorrencias", obj) if isinstance(obj, dict) else obj
    if not isinstance(itens, list):
        return ""
    linhas = []
    for o in itens[:40]:
        if isinstance(o, dict):
            cod = o.get("codigo") or o.get("id") or ""
            niv = o.get("nivel") or o.get("severidade") or ""
            tit = o.get("titulo") or o.get("descricao_curta") or ""
            linhas.append(f"- [{niv}] {cod}: {str(tit)[:120]}")
    return "\n".join(linhas)


# ── estado ──────────────────────────────────────────────────────

def _transicao_segura(audit: AuditIndex, cycle_id: str, destino: CycleState) -> None:
    """Transiciona apenas se válido a partir do estado atual; senão registra e ignora."""
    record = audit.get_cycle(cycle_id)
    if not record:
        return
    try:
        atual = CycleState(record["status"])
    except ValueError:
        logger.warning("Estado atual '%s' desconhecido — entrega não transiciona.", record["status"])
        return
    if atual == destino:
        return
    if destino in TRANSICOES_VALIDAS.get(atual, set()):
        audit.update_status(cycle_id, destino.value)
    else:
        logger.info("Transição %s → %s não permitida — entrega segue sem alterar estado.",
                    atual.value, destino.value)
