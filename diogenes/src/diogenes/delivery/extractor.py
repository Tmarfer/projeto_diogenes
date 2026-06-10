"""
delivery/extractor.py — DVA-CBS | Projeto Diógenes

ExtractorFinanceiro — leitura DETERMINÍSTICA dos números da planilha principal
do módulo via openpyxl. O LLM (Mycroft.mapear_dados_modulo) apenas localiza QUAIS
abas/células/intervalos correspondem a cada campo; aqui lemos o valor EXATO da
célula. Valor monetário nunca passa por transcrição de modelo — requisito de
auditoria do TC 015.848/2025-6.

Esquema do "mapa de extração" (produzido pelo agente, ver delivery/MAPA_EXTRACAO.md):

    {
      "planilha_principal": "AUX_MOD_10 PF - execução.xlsx",
      "blocos": [
        {"id": "visao_geral", "titulo": "Visão Geral", "descricao": "...",
         "kpis": [{"rotulo": "...", "aba": "Resumo", "celula": "B3",
                   "formato": "moeda|bilhoes|inteiro|percentual|texto",
                   "nota": "", "destaque": ""}],
         "tabelas": [{"titulo": "...", "aba": "Resumo", "intervalo": "A1:C10",
                      "cabecalho_na_primeira_linha": true, "fonte": "..."}]}
      ],
      "valores_agregados": [
        {"descricao": "Débitos totais", "aba": "Resumo",
         "celula_2023": "B12", "celula_2024": "C12", "formato": "bilhoes"}
      ],
      "sensibilidade_redutor": {"aba": "Sensibilidade",
        "intervalo_redutores": "A2:A6", "intervalo_valores": "B2:B6"}
    }
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from diogenes.models import (
    BlocoFinanceiro,
    CenarioSensibilidade,
    GraficoEntrega,
    KPIEntrega,
    MetodologiaCard,
    TabelaEntrega,
)

logger = logging.getLogger(__name__)


class ExtracaoError(Exception):
    """Falha ao abrir/ler a planilha principal do módulo."""


class ExtractorFinanceiro:
    def __init__(self, planilha_path: Path) -> None:
        self._path = Path(planilha_path)
        self._avisos: list[str] = []
        self._wb = None
        self._wb_formula = None

    @property
    def avisos(self) -> list[str]:
        return self._avisos

    def _abrir(self) -> None:
        if self._wb is not None:
            return
        try:
            import openpyxl
        except ImportError as exc:  # pragma: no cover
            raise ExtracaoError("openpyxl não disponível") from exc
        if not self._path.exists():
            raise ExtracaoError(f"Planilha principal não encontrada: {self._path}")
        # data_only=True → valores calculados (não fórmulas). Requer que a planilha
        # tenha sido salva por um editor que cacheia resultados (Excel/LibreOffice).
        # read_only=False: o acesso por intervalo (ws["A2:A6"]) é confiável neste modo;
        # planilhas de auditoria têm poucos MB, o que cabe em memória sem problema.
        self._wb = openpyxl.load_workbook(self._path, data_only=True)
        self._wb_formula = openpyxl.load_workbook(self._path, data_only=False)

    def extrair(self, mapa: dict[str, Any]) -> dict[str, Any]:
        """
        Aplica o mapa de extração e devolve as três estruturas financeiras do
        PacoteEntrega: blocos_financeiros, valores_agregados, sensibilidade_redutor.
        Nunca levanta por campo faltante — registra aviso e segue (best-effort),
        para que a entrega não trave por uma célula deslocada num único módulo.
        """
        if not mapa:
            return {"blocos_financeiros": [], "valores_agregados": [], "sensibilidade_redutor": None}
        self._abrir()

        blocos = [self._construir_bloco(b) for b in mapa.get("blocos", []) or []]

        agregados: list[dict[str, Any]] = []
        for va in mapa.get("valores_agregados", []) or []:
            fmt = va.get("formato", "bilhoes")
            agregados.append({
                "descricao": va.get("descricao", ""),
                "valor_2023": self._fmt(self._celula(va.get("aba"), va.get("celula_2023")), fmt),
                "valor_2024": self._fmt(self._celula(va.get("aba"), va.get("celula_2024")), fmt),
            })

        sensibilidade = None
        sm = mapa.get("sensibilidade_redutor")
        if isinstance(sm, dict):
            redutores = [self._num(v) for v in self._coluna(sm.get("aba"), sm.get("intervalo_redutores"))]
            valores = [self._num(v) for v in self._coluna(sm.get("aba"), sm.get("intervalo_valores"))]
            if redutores and valores:
                sensibilidade = {
                    "redutores": [int(r) for r in redutores if r is not None],
                    "valores": [float(v) for v in valores if v is not None],
                }

        return {
            "blocos_financeiros": blocos,
            "valores_agregados": agregados,
            "sensibilidade_redutor": sensibilidade,
        }

    # ── construção de blocos ─────────────────────────────────────

    def _construir_bloco(self, b: dict[str, Any]) -> BlocoFinanceiro:
        kpis: list[KPIEntrega] = []
        for k in b.get("kpis", []) or []:
            fmt = k.get("formato", "texto")
            bruto = self._celula(k.get("aba"), k.get("celula"))
            valor = self._fmt(bruto, fmt)
            delta, delta_tipo = self._delta(bruto, self._celula(k.get("aba"), k.get("celula_base")))
            kpis.append(KPIEntrega(
                rotulo=k.get("rotulo", ""),
                valor=valor,
                nota=k.get("nota", ""),
                destaque=k.get("destaque", ""),
                unidade=k.get("unidade", ""),
                delta=delta,
                delta_tipo=delta_tipo,
            ))

        cards = [
            MetodologiaCard(
                tag=c.get("tag", ""), titulo=c.get("titulo", ""),
                corpo=c.get("corpo", ""), chips=list(c.get("chips", []) or []),
            )
            for c in b.get("cards_metodologia", []) or []
        ]

        tabelas: list[TabelaEntrega] = []
        for t in b.get("tabelas", []) or []:
            matriz = self._matriz(t.get("aba"), t.get("intervalo"))
            headers: list[str] = []
            rows = matriz
            if t.get("cabecalho_na_primeira_linha", True) and matriz:
                headers = [str(c) if c is not None else "" for c in matriz[0]]
                rows = matriz[1:]
            rows_fmt = [[self._cell_str(c) for c in linha] for linha in rows]
            tabelas.append(TabelaEntrega(
                titulo=t.get("titulo", ""),
                subtitulo=t.get("subtitulo", ""),
                headers=headers,
                rows=rows_fmt,
                fonte=t.get("fonte", ""),
                total_labels=list(t.get("total_labels", []) or []),
            ))

        graficos: list[GraficoEntrega] = []
        for g in b.get("graficos", []) or []:
            labels = [self._cell_str(v) for v in self._coluna(g.get("aba"), g.get("intervalo_labels"))]
            series = [self._num(v) or 0.0 for v in self._coluna(g.get("aba"), g.get("intervalo_valores"))]
            graficos.append(GraficoEntrega(
                tipo=g.get("tipo", "barras"),
                titulo=g.get("titulo", ""),
                subtitulo=g.get("subtitulo", ""),
                layout=g.get("layout", "grid"),
                labels=labels,
                series=series,
                nota=g.get("nota", ""),
            ))

        cenarios: list[CenarioSensibilidade] = []
        for c in b.get("cenarios", []) or []:
            cenarios.append(CenarioSensibilidade(
                rotulo=c.get("rotulo", ""),
                valor=self._fmt(self._celula(c.get("aba"), c.get("celula")), c.get("formato", "texto")),
                nota=c.get("nota", ""),
                destaque=c.get("destaque", ""),
            ))

        return BlocoFinanceiro(
            id=b.get("id") or "",
            titulo=b.get("titulo") or "",
            descricao=b.get("descricao") or "",
            tipo=b.get("tipo") or self._inferir_tipo(b),
            narrativa=b.get("narrativa") or "",
            kpis=kpis,
            cards_metodologia=cards,
            tabelas=tabelas,
            graficos=graficos,
            cenarios=cenarios,
        )

    @staticmethod
    def _inferir_tipo(b: dict[str, Any]) -> str:
        if b.get("id") == "visao_geral" or b.get("cards_metodologia"):
            return "visao_geral"
        if b.get("cenarios"):
            return "sensibilidade"
        return "analitica"

    def _delta(self, atual: object, base: object) -> tuple[str, str]:
        """Variação percentual atual vs. base, lida deterministicamente das células."""
        na, nb = self._num(atual), self._num(base)
        if na is None or nb is None or nb == 0:
            return "", ""
        pct = (na - nb) / abs(nb) * 100
        tipo = "up" if pct > 0 else ("down" if pct < 0 else "neutral")
        sinal = "+" if pct > 0 else ""
        return f"{sinal}{_br_decimal(pct, 1)}%", tipo

    # ── leitura de células ───────────────────────────────────────

    def _ws(self, aba: str | None) -> Any:
        if not aba or self._wb is None:
            return None
        if aba in self._wb.sheetnames:
            return self._wb[aba]
        self._avisos.append(f"Aba '{aba}' ausente na planilha; campo ignorado.")
        return None

    def _celula(self, aba: str | None, ref: str | None) -> Any:
        if not ref:
            return None
        ws = self._ws(aba)
        if ws is None:
            return None
        try:
            val = ws[ref].value
            if val is None and self._wb_formula is not None and aba in self._wb_formula.sheetnames:
                # célula sem valor cacheado — registra a fórmula como aviso
                f = self._wb_formula[aba][ref].value
                if f is not None:
                    self._avisos.append(
                        f"Célula {aba}!{ref} sem valor calculado (fórmula '{f}'); "
                        f"reabra a planilha no Excel/LibreOffice e salve."
                    )
            return val
        except (KeyError, ValueError) as exc:
            self._avisos.append(f"Referência inválida {aba}!{ref}: {exc}")
            return None

    def _matriz(self, aba: str | None, intervalo: str | None) -> list[list[object]]:
        if not intervalo:
            return []
        ws = self._ws(aba)
        if ws is None:
            return []
        try:
            faixa = ws[intervalo]
        except (KeyError, ValueError) as exc:
            self._avisos.append(f"Intervalo inválido {aba}!{intervalo}: {exc}")
            return []
        # ws[intervalo] devolve uma tupla de linhas, cada uma com objetos Cell;
        # extraímos o .value para trabalhar com os valores, não as células.
        from openpyxl.cell.cell import Cell  # noqa: PLC0415
        if isinstance(faixa, Cell):          # referência de célula única (sem ':')
            return [[faixa.value]]
        return [[getattr(c, "value", c) for c in linha] for linha in faixa]

    def _coluna(self, aba: str | None, intervalo: str | None) -> list[object]:
        """Achata um intervalo (linha ou coluna) em uma lista simples."""
        return [c for linha in self._matriz(aba, intervalo) for c in linha]

    # ── formatação determinística ────────────────────────────────

    @staticmethod
    def _num(valor: object) -> float | None:
        if valor is None:
            return None
        if isinstance(valor, (int, float)):
            return float(valor)
        try:
            return float(str(valor).replace(".", "").replace(",", "."))
        except ValueError:
            return None

    @staticmethod
    def _cell_str(valor: object) -> str:
        if valor is None:
            return ""
        if isinstance(valor, float) and valor.is_integer():
            return _br_milhar(int(valor))
        if isinstance(valor, (int, float)):
            return _br_decimal(float(valor))
        return str(valor)

    def _fmt(self, valor: object, formato: str) -> str:
        n = self._num(valor)
        if n is None:
            return self._cell_str(valor)
        if formato == "moeda":
            return "R$ " + _br_decimal(n)
        if formato == "bilhoes":
            return f"R$ {_br_decimal(n / 1_000_000_000)} bilhões"
        if formato == "milhoes":
            return f"R$ {_br_decimal(n / 1_000_000)} milhões"
        if formato == "inteiro":
            return _br_milhar(int(round(n)))
        if formato == "percentual":
            return _br_decimal(n) + "%"
        return self._cell_str(valor)


def _br_milhar(n: int) -> str:
    return f"{n:,}".replace(",", ".")


def _br_decimal(n: float, casas: int = 2) -> str:
    s = f"{n:,.{casas}f}"                       # 1,234,567.89
    return s.replace(",", "§").replace(".", ",").replace("§", ".")
