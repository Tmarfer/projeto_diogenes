"""
motors/motor_perfilamento.py — DVA-CBS | Projeto Diógenes

MotorPerfilamento: análise estatística determinística de CSV/XLSX via DuckDB.
Roda após motor_start, antes de Irene/Watson.

Por cada arquivo analisável:
  - CSV  → DuckDB (read_csv nativo): row_count, por coluna: null_rate,
            distinct_count, min/max, unicidade em colunas-chave candidatas.
  - XLSX → openpyxl read_only + stats manuais (DuckDB não lê XLSX nativamente).
  Grava: {cycle_dir}/_runtime/perfis/{safe_filename}.perfil.yaml

Cross-file:
  Detecta colunas compartilhadas entre arquivos (candidatas a join/relação).
  Grava: {cycle_dir}/_runtime/perfil_pacote.yaml

Não usa LLM. Best-effort: erros em arquivos individuais nunca abortam o ciclo.
Referência: Proposta de Solução CSV/XLSX — Auditoria MOD_010_A1_20260605T103922Z.
"""
from __future__ import annotations

import hashlib
import logging
import re
from dataclasses import asdict, dataclass, field
from datetime import UTC, datetime
from pathlib import Path

import yaml

from diogenes.agents.file_prep import classificar

logger = logging.getLogger(__name__)

# ── Constantes de configuração ─────────────────────────────────────────────

# Extensões com análise estatística via motor (subconjunto de EXTENSOES_ANALISE)
_EXTENSOES_PERFILAMENTO = frozenset({".csv", ".xlsx"})

# Palavras-chave para inferir colunas-chave candidatas (comparação case-insensitive)
_KEYWORDS_CHAVE = (
    "cpf", "cnpj", "id", "chave", "codigo", "código",
    "matricula", "matrícula", "inscricao", "inscrição",
    "nfe", "nota_fiscal", "contrato", "protocolo", "seq",
)
# Limiar de nulos para alertas
_NULO_LIMIAR_MEDIA = 0.05    # ≥ 5%  → MEDIA
_NULO_LIMIAR_ALTA  = 0.20    # ≥ 20% → ALTA
# Limiar de unicidade para alertas em colunas-chave (1.0 = sem duplicatas)
_UNICIDADE_MINIMA = 0.9990   # < 99,9% → DUPLICATA_CHAVE
# Máximo de linhas por aba XLSX (evita OOM em arquivos gigantes)
_XLSX_MAX_ROWS = 100_000
# Nº de valores duplicados incluídos no alerta (mascarados para PII)
_AMOSTRA_DUPLICATAS = 5


# ── Dataclasses de perfil ──────────────────────────────────────────────────

@dataclass
class ColunaPerfil:
    nome: str
    tipo_inferido: str     # "texto_id" | "numerico" | "categorico" | "texto"
    row_count: int
    null_count: int
    null_pct: float
    distinct_count: int
    unicidade_pct: float
    eh_chave_candidata: bool = False
    alerta_duplicatas: bool = False
    duplicatas_count: int = 0
    min_valor: str | None = None
    max_valor: str | None = None
    media_valor: str | None = None
    valores_amostra: list[str] = field(default_factory=list)


@dataclass
class AlertaIntegridade:
    tipo: str              # "DUPLICATA_CHAVE" | "NULOS_SIGNIFICATIVOS" | "ARQUIVO_GRANDE"
    coluna: str | None
    contagem: int
    percentual: float
    severidade: str        # "CRITICA" | "ALTA" | "MEDIA" | "BAIXA"
    detalhe: str = ""


@dataclass
class PerfilArquivo:
    arquivo: str
    caminho_relativo: str
    hash_sha256: str
    gerado_em: str
    row_count: int
    col_count: int
    colunas: list[ColunaPerfil]
    alertas: list[AlertaIntegridade]
    abas: list[str] = field(default_factory=list)
    truncado: bool = False     # True se XLSX foi cortado em _XLSX_MAX_ROWS
    erro: str | None = None


@dataclass
class PerfilPacote:
    gerado_em: str
    inputs_dir: str
    total_arquivos: int
    total_alertas_criticos: int
    perfis_dir: str
    arquivos_perfilados: list[str]
    arquivos_com_erro: list[str]
    colunas_compartilhadas: list[dict] = field(default_factory=list)


# ── API pública ────────────────────────────────────────────────────────────

def perfilar_diretorio(inputs_dir: Path, perfis_dir: Path) -> PerfilPacote:
    """
    Perfila todos os CSV e XLSX analisáveis em inputs_dir.
    Grava um .perfil.yaml por arquivo e um perfil_pacote.yaml de sumário.
    Nunca levanta — erros por arquivo são registrados em perfil.erro.
    """
    perfis_dir.mkdir(parents=True, exist_ok=True)
    agora = datetime.now(UTC).isoformat()

    perfilados: list[str] = []
    com_erro: list[str] = []
    criticos_total = 0
    todos_colunas: dict[str, list[str]] = {}   # nome_col → [arquivos que a têm]

    for path in sorted(inputs_dir.rglob("*")):
        if not path.is_file():
            continue
        if path.suffix.lower() not in _EXTENSOES_PERFILAMENTO:
            continue
        rel = path.relative_to(inputs_dir)
        analisavel, _ = classificar(rel)
        if not analisavel:
            continue

        perfil = perfilar_arquivo(path, rel)
        _salvar_perfil(perfil, perfis_dir)

        if perfil.erro:
            com_erro.append(str(rel))
        else:
            perfilados.append(str(rel))
            criticos_total += sum(
                1 for a in perfil.alertas if a.severidade == "CRITICA"
            )
            for col in perfil.colunas:
                todos_colunas.setdefault(col.nome, []).append(str(rel))

    # Cross-file: colunas presentes em ≥ 2 arquivos são candidatas a join
    compartilhadas = [
        {"coluna": col, "arquivos": arqs}
        for col, arqs in sorted(todos_colunas.items())
        if len(arqs) >= 2
    ]

    pacote = PerfilPacote(
        gerado_em=agora,
        inputs_dir=str(inputs_dir),
        total_arquivos=len(perfilados) + len(com_erro),
        total_alertas_criticos=criticos_total,
        perfis_dir=str(perfis_dir),
        arquivos_perfilados=perfilados,
        arquivos_com_erro=com_erro,
        colunas_compartilhadas=compartilhadas,
    )
    _salvar_pacote(pacote, perfis_dir)
    logger.info(
        "[Motor Perfilamento] %d arquivos perfilados, %d erros, %d alertas críticos",
        len(perfilados), len(com_erro), criticos_total,
    )
    return pacote


def perfilar_arquivo(path: Path, rel_path: Path | None = None) -> PerfilArquivo:
    """
    Perfila um único CSV ou XLSX. Nunca levanta — captura e registra em .erro.
    """
    agora = datetime.now(UTC).isoformat()
    rel_str = str(rel_path) if rel_path else path.name
    h256 = _hash_arquivo(path)

    try:
        ext = path.suffix.lower()
        if ext == ".csv":
            row_count, col_count, colunas, alertas = _perfilar_csv(path)
            return PerfilArquivo(
                arquivo=path.name, caminho_relativo=rel_str,
                hash_sha256=h256, gerado_em=agora,
                row_count=row_count, col_count=col_count,
                colunas=colunas, alertas=alertas,
            )
        if ext == ".xlsx":
            row_count, col_count, colunas, alertas, abas, truncado = _perfilar_xlsx(path)
            return PerfilArquivo(
                arquivo=path.name, caminho_relativo=rel_str,
                hash_sha256=h256, gerado_em=agora,
                row_count=row_count, col_count=col_count,
                colunas=colunas, alertas=alertas,
                abas=abas, truncado=truncado,
            )
    except Exception as exc:  # noqa: BLE001
        logger.warning("[Motor Perfilamento] Erro ao perfilar %s: %s", path, exc)
        return PerfilArquivo(
            arquivo=path.name, caminho_relativo=rel_str,
            hash_sha256=h256, gerado_em=agora,
            row_count=0, col_count=0,
            colunas=[], alertas=[],
            erro=str(exc),
        )
    return PerfilArquivo(  # extensão não suportada (não deveria chegar aqui)
        arquivo=path.name, caminho_relativo=rel_str,
        hash_sha256=h256, gerado_em=agora,
        row_count=0, col_count=0,
        colunas=[], alertas=[],
        erro="Extensão não suportada pelo motor de perfilamento.",
    )


def caminho_perfil(perfis_dir: Path, arquivo_path: Path) -> Path:
    """Caminho canônico do .perfil.yaml para um arquivo."""
    safe = re.sub(r"[^A-Za-z0-9_.\-]", "_", arquivo_path.name)
    return perfis_dir / f"{safe}.perfil.yaml"


def ler_perfil_texto(perfis_dir: Path, arquivo_path: Path) -> str:
    """
    Lê o .perfil.yaml como texto formatado para injeção em prompt Watson.
    Retorna "" se o arquivo não existe (motor não rodou ou arquivo novo).
    """
    p = caminho_perfil(perfis_dir, arquivo_path)
    if not p.exists():
        return ""
    try:
        return p.read_text(encoding="utf-8")
    except Exception as exc:  # noqa: BLE001
        logger.warning("[Motor Perfilamento] Erro ao ler perfil %s: %s", p, exc)
        return ""


# ── CSV via DuckDB ─────────────────────────────────────────────────────────

def _perfilar_csv(
    path: Path,
) -> tuple[int, int, list[ColunaPerfil], list[AlertaIntegridade]]:
    """Analisa CSV via DuckDB. Retorna (row_count, col_count, colunas, alertas)."""
    import duckdb  # importação local: dep opcional que pode não estar instalada

    conn = duckdb.connect()
    try:
        # Usa read_csv com auto_detect — tolera separadores variados e encodings.
        # ignore_errors=true evita abort por linhas malformadas.
        conn.execute(
            "CREATE TABLE t AS SELECT * FROM read_csv(?, "
            "header=True, auto_detect=True, ignore_errors=true)",
            [str(path)],
        )
        row_count: int = conn.execute("SELECT COUNT(*) FROM t").fetchone()[0]  # type: ignore[index]
        cols_info = conn.execute("DESCRIBE t").fetchall()
        # cols_info: [(nome, tipo, null, key, default, extra), ...]
        col_count = len(cols_info)

        colunas: list[ColunaPerfil] = []
        for row in cols_info:
            col_name: str = row[0]
            col_type: str = row[1]
            colunas.append(_stats_coluna_duckdb(conn, "t", col_name, col_type, row_count))

        conn.close()
        alertas = _gerar_alertas(colunas)
        return row_count, col_count, colunas, alertas
    except Exception:
        conn.close()
        raise


def _stats_coluna_duckdb(
    conn: object, tabela: str, nome: str, tipo: str, total_linhas: int
) -> ColunaPerfil:
    """Computa estatísticas de uma coluna DuckDB."""
    import duckdb  # noqa: F401 — importado para type hint abaixo

    # Escapar nome da coluna com aspas duplas
    col_q = f'"{nome}"'

    null_count: int = conn.execute(  # type: ignore[union-attr]
        f"SELECT COUNT(*) FROM {tabela} WHERE {col_q} IS NULL"
    ).fetchone()[0]  # type: ignore[index]

    distinct_count: int = conn.execute(  # type: ignore[union-attr]
        f"SELECT COUNT(DISTINCT {col_q}) FROM {tabela}"
    ).fetchone()[0]  # type: ignore[index]

    null_pct = null_count / total_linhas if total_linhas > 0 else 0.0
    unicidade_pct = distinct_count / total_linhas if total_linhas > 0 else 1.0
    eh_chave = _eh_chave_candidata(nome)

    tipo_inferido = _inferir_tipo(tipo, eh_chave, distinct_count, total_linhas)

    # Estatísticas de range apenas para tipos numéricos
    min_v = max_v = media_v = None
    _TIPOS_NUMERICOS = {
        "BIGINT", "INTEGER", "DOUBLE", "FLOAT", "DECIMAL",
        "HUGEINT", "SMALLINT", "TINYINT", "INT", "UBIGINT",
    }
    tipo_upper = tipo.upper().split("(")[0].strip()
    if tipo_upper in _TIPOS_NUMERICOS:
        try:
            row = conn.execute(  # type: ignore[union-attr]
                f"SELECT MIN({col_q}), MAX({col_q}), AVG(CAST({col_q} AS DOUBLE)) FROM {tabela}"
            ).fetchone()
            if row:
                min_v = str(row[0]) if row[0] is not None else None
                max_v = str(row[1]) if row[1] is not None else None
                media_v = f"{row[2]:.4f}" if row[2] is not None else None
        except Exception:  # noqa: BLE001
            pass

    # Amostrar duplicatas em colunas-chave
    duplicatas_count = 0
    alerta_dup = False
    if eh_chave and unicidade_pct < _UNICIDADE_MINIMA and total_linhas > 0:
        alerta_dup = True
        duplicatas_count = total_linhas - distinct_count

    # Amostra de valores para contexto (top 3 non-null, apenas para categóricos)
    valores_amostra: list[str] = []
    if tipo_inferido == "categorico" and distinct_count <= 20:
        try:
            rows = conn.execute(  # type: ignore[union-attr]
                f"SELECT DISTINCT {col_q} FROM {tabela} "
                f"WHERE {col_q} IS NOT NULL LIMIT 5"
            ).fetchall()
            valores_amostra = [str(r[0]) for r in rows]
        except Exception:  # noqa: BLE001
            pass

    return ColunaPerfil(
        nome=nome, tipo_inferido=tipo_inferido,
        row_count=total_linhas, null_count=null_count, null_pct=round(null_pct, 4),
        distinct_count=distinct_count, unicidade_pct=round(unicidade_pct, 6),
        eh_chave_candidata=eh_chave, alerta_duplicatas=alerta_dup,
        duplicatas_count=duplicatas_count,
        min_valor=min_v, max_valor=max_v, media_valor=media_v,
        valores_amostra=valores_amostra,
    )


# ── XLSX via openpyxl ──────────────────────────────────────────────────────

def _perfilar_xlsx(
    path: Path,
) -> tuple[int, int, list[ColunaPerfil], list[AlertaIntegridade], list[str], bool]:
    """
    Analisa XLSX via openpyxl (read_only).
    Analisa apenas a primeira aba (a mais relevante para integridade numérica).
    Retorna (row_count, col_count, colunas, alertas, lista_abas, truncado).
    """
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    abas = list(wb.sheetnames)
    if not abas:
        wb.close()
        return 0, 0, [], [], [], False

    # Analisa a primeira aba
    ws = wb[abas[0]]
    header: list[str] = []
    col_values: dict[str, list[object]] = {}
    truncado = False
    linhas_lidas = 0

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header = [
                str(c).strip() if c is not None else f"col_{j}"
                for j, c in enumerate(row)
            ]
            col_values = {h: [] for h in header}
            continue
        if i > _XLSX_MAX_ROWS:
            truncado = True
            break
        for h, val in zip(header, row, strict=False):
            col_values[h].append(val)
        linhas_lidas += 1

    wb.close()

    row_count = linhas_lidas
    col_count = len(header)
    colunas: list[ColunaPerfil] = [
        _stats_coluna_lista(col_values.get(h, []), h, row_count)
        for h in header
    ]
    alertas = _gerar_alertas(colunas)
    return row_count, col_count, colunas, alertas, abas, truncado


def _stats_coluna_lista(
    valores: list[object], nome: str, total_linhas: int
) -> ColunaPerfil:
    """Computa estatísticas de uma coluna a partir de uma lista Python."""
    null_count = sum(1 for v in valores if v is None or (isinstance(v, str) and v.strip() == ""))
    nao_nulos = [v for v in valores if v is not None and not (isinstance(v, str) and v.strip() == "")]
    distinct_count = len(set(str(v) for v in nao_nulos))
    null_pct = null_count / total_linhas if total_linhas > 0 else 0.0
    unicidade_pct = distinct_count / total_linhas if total_linhas > 0 else 1.0
    eh_chave = _eh_chave_candidata(nome)

    # Inferência de tipo pelo conteúdo
    numericos = _extrair_numericos(nao_nulos)
    if len(numericos) >= max(1, len(nao_nulos) * 0.8):
        tipo_inferido = "numerico"
    else:
        tipo_inferido = _inferir_tipo_lista(nome, distinct_count, total_linhas, eh_chave)

    min_v = max_v = media_v = None
    if tipo_inferido == "numerico" and numericos:
        min_v = str(min(numericos))
        max_v = str(max(numericos))
        media_v = f"{sum(numericos) / len(numericos):.4f}"

    alerta_dup = eh_chave and unicidade_pct < _UNICIDADE_MINIMA and total_linhas > 0
    duplicatas_count = max(0, total_linhas - distinct_count - null_count) if alerta_dup else 0

    valores_amostra: list[str] = []
    if tipo_inferido == "categorico" and distinct_count <= 20 and nao_nulos:
        valores_amostra = list({str(v) for v in nao_nulos})[:5]

    return ColunaPerfil(
        nome=nome, tipo_inferido=tipo_inferido,
        row_count=total_linhas, null_count=null_count, null_pct=round(null_pct, 4),
        distinct_count=distinct_count, unicidade_pct=round(unicidade_pct, 6),
        eh_chave_candidata=eh_chave, alerta_duplicatas=alerta_dup,
        duplicatas_count=duplicatas_count,
        min_valor=min_v, max_valor=max_v, media_valor=media_v,
        valores_amostra=valores_amostra,
    )


# ── Lógica de classificação e alertas ─────────────────────────────────────

def _eh_chave_candidata(nome: str) -> bool:
    """True se o nome da coluna sugere uma chave de negócio."""
    nome_lower = nome.lower()
    return any(k in nome_lower for k in _KEYWORDS_CHAVE)


def _inferir_tipo(
    duckdb_tipo: str, eh_chave: bool, distinct_count: int, total: int
) -> str:
    tipo = duckdb_tipo.upper().split("(")[0].strip()
    _NUMERICOS = {
        "BIGINT", "INTEGER", "DOUBLE", "FLOAT", "DECIMAL",
        "HUGEINT", "SMALLINT", "TINYINT", "INT", "UBIGINT",
    }
    if tipo in _NUMERICOS:
        return "numerico"
    if eh_chave:
        return "texto_id"
    if total > 0 and distinct_count / total < 0.05:
        return "categorico"
    return "texto"


def _inferir_tipo_lista(
    nome: str, distinct_count: int, total: int, eh_chave: bool
) -> str:
    if eh_chave:
        return "texto_id"
    if total > 0 and distinct_count / total < 0.05:
        return "categorico"
    return "texto"


def _extrair_numericos(valores: list[object]) -> list[float]:
    """Extrai floats de uma lista heterogênea."""
    result: list[float] = []
    for v in valores:
        if isinstance(v, (int, float)):
            result.append(float(v))
        elif isinstance(v, str):
            import contextlib
            with contextlib.suppress(ValueError):
                result.append(float(v.replace(",", ".")))
    return result


def _gerar_alertas(colunas: list[ColunaPerfil]) -> list[AlertaIntegridade]:
    """Gera alertas de integridade a partir das estatísticas das colunas."""
    alertas: list[AlertaIntegridade] = []
    for col in colunas:
        # Duplicatas em coluna-chave
        if col.alerta_duplicatas and col.duplicatas_count > 0:
            severidade = "CRITICA" if col.eh_chave_candidata else "ALTA"
            alertas.append(AlertaIntegridade(
                tipo="DUPLICATA_CHAVE",
                coluna=col.nome,
                contagem=col.duplicatas_count,
                percentual=round((1.0 - col.unicidade_pct) * 100, 2),
                severidade=severidade,
                detalhe=(
                    f"Coluna '{col.nome}' tem {col.duplicatas_count} registros duplicados "
                    f"({(1 - col.unicidade_pct) * 100:.2f}% do total). "
                    f"Unicidade: {col.unicidade_pct * 100:.4f}%."
                ),
            ))
        # Nulos significativos
        if col.null_pct >= _NULO_LIMIAR_ALTA:
            alertas.append(AlertaIntegridade(
                tipo="NULOS_SIGNIFICATIVOS",
                coluna=col.nome,
                contagem=col.null_count,
                percentual=round(col.null_pct * 100, 2),
                severidade="ALTA",
                detalhe=f"Coluna '{col.nome}': {col.null_pct * 100:.1f}% nulos.",
            ))
        elif col.null_pct >= _NULO_LIMIAR_MEDIA:
            alertas.append(AlertaIntegridade(
                tipo="NULOS_SIGNIFICATIVOS",
                coluna=col.nome,
                contagem=col.null_count,
                percentual=round(col.null_pct * 100, 2),
                severidade="MEDIA",
                detalhe=f"Coluna '{col.nome}': {col.null_pct * 100:.1f}% nulos.",
            ))
    return alertas


# ── Persistência ───────────────────────────────────────────────────────────

def _salvar_perfil(perfil: PerfilArquivo, perfis_dir: Path) -> None:
    """Serializa PerfilArquivo para YAML e grava em perfis_dir."""
    dest = caminho_perfil(perfis_dir, Path(perfil.arquivo))
    dados = asdict(perfil)
    try:
        dest.write_text(
            yaml.dump(dados, allow_unicode=True, default_flow_style=False, sort_keys=False),
            encoding="utf-8",
        )
    except Exception as exc:  # noqa: BLE001
        logger.warning("[Motor Perfilamento] Falha ao gravar perfil %s: %s", dest, exc)


def _salvar_pacote(pacote: PerfilPacote, perfis_dir: Path) -> None:
    """Serializa PerfilPacote para YAML e grava como perfil_pacote.yaml."""
    dest = perfis_dir / "perfil_pacote.yaml"
    dados = asdict(pacote)
    try:
        dest.write_text(
            yaml.dump(dados, allow_unicode=True, default_flow_style=False, sort_keys=False),
            encoding="utf-8",
        )
    except Exception as exc:  # noqa: BLE001
        logger.warning("[Motor Perfilamento] Falha ao gravar pacote %s: %s", dest, exc)


# ── Utilitários ────────────────────────────────────────────────────────────

def _hash_arquivo(path: Path) -> str:
    """SHA-256 do arquivo. Retorna string vazia em caso de erro (best-effort)."""
    try:
        h = hashlib.sha256()
        with path.open("rb") as f:
            for chunk in iter(lambda: f.read(65536), b""):
                h.update(chunk)
        return h.hexdigest()
    except Exception:  # noqa: BLE001
        return ""
