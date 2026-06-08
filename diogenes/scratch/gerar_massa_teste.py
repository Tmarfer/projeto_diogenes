"""
scratch/gerar_massa_teste.py — DVA-CBS | Projeto Diógenes
Gera massa de teste MOD_SINT_001 (~30 arquivos) para teste ponta-a-ponta.

Simula pacote RFB/CBS com: XLSX, CSV, SQL, TXT, MD.
Execute UMA VEZ da raiz do projeto:
    python scratch/gerar_massa_teste.py
"""
from __future__ import annotations

import csv
import textwrap
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Respeita DIOGENES_WORKSPACE se definida; fallback ao ./workspace local
import os
_ws_env = os.environ.get("DIOGENES_WORKSPACE")
if _ws_env:
    OUT = Path(_ws_env).expanduser().resolve() / "input" / "MOD_SINT_001"
else:
    OUT = Path(__file__).parent.parent / "workspace" / "input" / "MOD_SINT_001"
OUT.mkdir(parents=True, exist_ok=True)

# ── Paleta de cabeçalho ───────────────────────────────────────────────────────
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _wb_header(ws, cols: list[str]) -> None:
    """Aplica estilo ao cabeçalho da primeira linha."""
    for i, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i, value=col)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = HDR_ALIGN
        ws.column_dimensions[get_column_letter(i)].width = max(len(col) + 4, 14)


def _save_xlsx_csv(filename_base: str, cols: list[str], rows: list[list]) -> None:
    """Salva <filename_base>.xlsx e <filename_base>.csv com os mesmos dados."""
    # XLSX
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = filename_base[:31]
    _wb_header(ws, cols)
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    ws.freeze_panes = "A2"
    wb.save(OUT / f"{filename_base}.xlsx")

    # CSV
    with open(OUT / f"{filename_base}.csv", "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(cols)
        writer.writerows(rows)

    print(f"  ✓ {filename_base}.xlsx + .csv")


# ══════════════════════════════════════════════════════════════════════════════
# 1. RECEITA BRUTA — Empresa A (CNPJ 12.345.678/0001-90)
# ══════════════════════════════════════════════════════════════════════════════
cols_rb = [
    "CNPJ", "Competência", "Descrição Operação", "CFOP",
    "Código NCM", "Valor Bruto (R$)", "Desconto (R$)", "Valor Líquido (R$)",
    "Base Cálculo CBS (R$)", "Observação",
]
rows_rb_a = [
    ["12.345.678/0001-90", "01/2025", "Venda de mercadorias nacionais", "5102", "7308.90.90",
     850_000.00, 12_500.00, 837_500.00, 837_500.00, "Produto siderúrgico — alíquota cheia"],
    ["12.345.678/0001-90", "02/2025", "Prestação de serviços TI", "5933", "8471.30.19",
     320_000.00, 0.00, 320_000.00, 320_000.00, "Software sob encomenda — LC 214/2025 Art. 24"],
    ["12.345.678/0001-90", "03/2025", "Venda de mercadorias exportação", "7101", "7308.90.90",
     420_000.00, 5_000.00, 415_000.00, 0.00, "Exportação — imune CBS (LC 214 Art. 12 §1°)"],
    ["12.345.678/0001-90", "04/2025", "Receita financeira", "0000", "—",
     15_000.00, 0.00, 15_000.00, 15_000.00, "Juros recebidos — CBS sobre rend. financeiros"],
    ["12.345.678/0001-90", "05/2025", "Venda imobilizado", "5551", "8474.20.00",
     90_000.00, 0.00, 90_000.00, 90_000.00, "Ativo fixo — tributação integral"],
    ["12.345.678/0001-90", "06/2025", "Venda de mercadorias nacionais", "5102", "2710.12.59",
     1_200_000.00, 22_000.00, 1_178_000.00, 1_178_000.00, "Combustível — redução setorial LC 214 Art. 39"],
]
_save_xlsx_csv("receita_bruta_empresa_A", cols_rb, rows_rb_a)

# ══════════════════════════════════════════════════════════════════════════════
# 2. RECEITA BRUTA — Empresa B (CNPJ 98.765.432/0001-10)
# ══════════════════════════════════════════════════════════════════════════════
rows_rb_b = [
    ["98.765.432/0001-10", "01/2025", "Venda de alimentos básicos", "5102", "1006.30.21",
     540_000.00, 8_000.00, 532_000.00, 0.00, "Cesta básica — alíquota zero CBS Art. 47"],
    ["98.765.432/0001-10", "02/2025", "Venda de medicamentos", "5102", "3003.90.89",
     280_000.00, 0.00, 280_000.00, 0.00, "Med. essencial — isenção CBS LC 214 Anexo VII"],
    ["98.765.432/0001-10", "03/2025", "Serviços de saúde", "5933", "8620.00.00",
     190_000.00, 0.00, 190_000.00, 95_000.00, "Serv. saúde — alíquota reduzida 50%"],
    ["98.765.432/0001-10", "04/2025", "Venda de hortifrúti", "5102", "0702.00.00",
     220_000.00, 4_500.00, 215_500.00, 0.00, "Produto agropecuário — alíquota zero"],
    ["98.765.432/0001-10", "05/2025", "Receita de locação de bem móvel", "0000", "—",
     45_000.00, 0.00, 45_000.00, 45_000.00, "Locação — art. 71 LC 214"],
    ["98.765.432/0001-10", "06/2025", "Venda de ovos e laticínios", "5102", "0401.10.10",
     175_000.00, 2_200.00, 172_800.00, 0.00, "Alíquota zero — lista alimentos básicos"],
]
_save_xlsx_csv("receita_bruta_empresa_B", cols_rb, rows_rb_b)

# ══════════════════════════════════════════════════════════════════════════════
# 3. NF-e EMITIDAS — Empresa A
# ══════════════════════════════════════════════════════════════════════════════
cols_nfe = [
    "Chave NF-e", "Data Emissão", "CNPJ Emitente", "CNPJ Destinatário",
    "Valor Total (R$)", "Base CBS (R$)", "Alíquota CBS (%)", "CBS Destacado (R$)",
    "CFOP", "NCM Principal", "Operação",
]
rows_nfe_a = [
    ["3525010112345678000190550010000001011234567890", "2025-01-15",
     "12.345.678/0001-90", "11.111.111/0001-11",
     245_000.00, 245_000.00, 9.9, 24_255.00, "5102", "7308.90.90", "Venda nacional"],
    ["3525020112345678000190550010000002021234567891", "2025-02-18",
     "12.345.678/0001-90", "22.222.222/0001-22",
     98_500.00, 98_500.00, 9.9, 9_751.50, "5102", "7308.90.90", "Venda nacional"],
    ["3525030112345678000190550010000003031234567892", "2025-03-22",
     "12.345.678/0001-90", "33.333.333/0001-33",
     415_000.00, 0.00, 0.0, 0.00, "7101", "7308.90.90", "Exportação"],
    ["3525040112345678000190550010000004041234567893", "2025-04-10",
     "12.345.678/0001-90", "44.444.444/0001-44",
     320_000.00, 320_000.00, 9.9, 31_680.00, "5933", "8471.30.19", "Serviço TI"],
    ["3525050112345678000190550010000005051234567894", "2025-05-05",
     "12.345.678/0001-90", "55.555.555/0001-55",
     90_000.00, 90_000.00, 9.9, 8_910.00, "5551", "8474.20.00", "Imobilizado"],
]
_save_xlsx_csv("nfe_emitidas_empresa_A", cols_nfe, rows_nfe_a)

# ══════════════════════════════════════════════════════════════════════════════
# 4. NF-e EMITIDAS — Empresa B
# ══════════════════════════════════════════════════════════════════════════════
rows_nfe_b = [
    ["3525010198765432000110550010000001011234567899", "2025-01-08",
     "98.765.432/0001-10", "66.666.666/0001-66",
     180_000.00, 0.00, 0.0, 0.00, "5102", "1006.30.21", "Alimento básico — alíq. zero"],
    ["3525020198765432000110550010000002021234567898", "2025-02-12",
     "98.765.432/0001-10", "77.777.777/0001-77",
     92_000.00, 0.00, 0.0, 0.00, "5102", "3003.90.89", "Medicamento — isento"],
    ["3525030198765432000110550010000003031234567897", "2025-03-19",
     "98.765.432/0001-10", "88.888.888/0001-88",
     190_000.00, 95_000.00, 4.95, 9_402.50, "5933", "8620.00.00", "Serv. saúde — alíq. reduzida"],
    ["3525040198765432000110550010000004041234567896", "2025-04-25",
     "98.765.432/0001-10", "99.999.999/0001-99",
     45_000.00, 45_000.00, 9.9, 4_455.00, "0000", "—", "Locação bem móvel"],
]
_save_xlsx_csv("nfe_emitidas_empresa_B", cols_nfe, rows_nfe_b)

# ══════════════════════════════════════════════════════════════════════════════
# 5. CRÉDITOS PIS/COFINS → CBS (créditos acumulados pré-transição)
# ══════════════════════════════════════════════════════════════════════════════
cols_cred = [
    "CNPJ", "Tipo Crédito", "Referência", "Valor Crédito PIS (R$)",
    "Valor Crédito COFINS (R$)", "Total Crédito (R$)",
    "Aproveitado (R$)", "Saldo (R$)", "Base Legal",
]
rows_cred = [
    ["12.345.678/0001-90", "Aquisição insumo", "dez/2024",
     8_500.00, 39_100.00, 47_600.00, 47_600.00, 0.00, "LC 214 Art. 54 §2°"],
    ["12.345.678/0001-90", "Energia elétrica", "dez/2024",
     1_200.00, 5_520.00, 6_720.00, 6_720.00, 0.00, "LC 214 Art. 54 §3° I"],
    ["12.345.678/0001-90", "Serv. transporte", "jan/2025",
     2_100.00, 9_660.00, 11_760.00, 5_000.00, 6_760.00, "LC 214 Art. 54 §3° II"],
    ["98.765.432/0001-10", "Aquisição insumo alimentar", "dez/2024",
     4_200.00, 19_320.00, 23_520.00, 23_520.00, 0.00, "LC 214 Art. 54 §5°"],
    ["98.765.432/0001-10", "Arrendamento máquinas", "jan/2025",
     900.00, 4_140.00, 5_040.00, 2_500.00, 2_540.00, "LC 214 Art. 54 §4°"],
]
_save_xlsx_csv("creditos_pis_cofins", cols_cred, rows_cred)

# ══════════════════════════════════════════════════════════════════════════════
# 6. REDUÇÕES SETORIAIS
# ══════════════════════════════════════════════════════════════════════════════
cols_red = [
    "CNPJ", "Setor Econômico", "Dispositivo LC 214", "Alíquota Cheia CBS (%)",
    "Fator Redução (%)", "Alíquota Efetiva (%)", "Base Cálculo (R$)",
    "CBS sem Redução (R$)", "CBS com Redução (R$)", "Diferença (R$)",
]
rows_red = [
    ["12.345.678/0001-90", "Combustíveis derivados petróleo", "Art. 39",
     9.9, 44.0, 5.544, 1_178_000.00, 116_622.00, 65_288.32, 51_333.68],
    ["98.765.432/0001-10", "Serviços de saúde humana", "Art. 44",
     9.9, 50.0, 4.95, 190_000.00, 18_810.00, 9_405.00, 9_405.00],
    ["98.765.432/0001-10", "Educação básica", "Art. 45",
     9.9, 60.0, 3.96, 85_000.00, 8_415.00, 3_366.00, 5_049.00],
    ["12.345.678/0001-90", "Transporte coletivo urbano", "Art. 41",
     9.9, 100.0, 0.0, 45_000.00, 4_455.00, 0.00, 4_455.00],
]
_save_xlsx_csv("reducoes_setoriais", cols_red, rows_red)

# ══════════════════════════════════════════════════════════════════════════════
# 7. MEMÓRIA DE CÁLCULO CBS — consolidada
# ══════════════════════════════════════════════════════════════════════════════
cols_mc = [
    "CNPJ", "Competência", "Base Cálculo Total (R$)", "Exclusões (R$)",
    "Reduções Setoriais (R$)", "Base CBS Líquida (R$)", "Alíquota Aplicada (%)",
    "CBS Bruta (R$)", "Créditos CBS (R$)", "CBS a Recolher (R$)",
    "DARF Recolhido (R$)", "Diferença Apurada (R$)",
]
rows_mc = [
    ["12.345.678/0001-90", "01/2025",
     850_000.00, 0.00, 0.00, 850_000.00, 9.9, 84_150.00, 47_600.00, 36_550.00, 36_550.00, 0.00],
    ["12.345.678/0001-90", "02/2025",
     320_000.00, 0.00, 0.00, 320_000.00, 9.9, 31_680.00, 11_760.00, 19_920.00, 19_920.00, 0.00],
    ["12.345.678/0001-90", "03/2025",
     415_000.00, 415_000.00, 0.00, 0.00, 0.0, 0.00, 0.00, 0.00, 0.00, 0.00],
    ["12.345.678/0001-90", "04/2025",
     15_000.00, 0.00, 0.00, 15_000.00, 9.9, 1_485.00, 0.00, 1_485.00, 1_485.00, 0.00],
    ["12.345.678/0001-90", "05/2025",
     90_000.00, 0.00, 0.00, 90_000.00, 9.9, 8_910.00, 0.00, 8_910.00, 8_910.00, 0.00],
    ["12.345.678/0001-90", "06/2025",
     1_178_000.00, 0.00, 51_333.68, 1_126_666.32, 9.9, 111_539.97, 0.00, 111_539.97, 111_539.97, 0.00],
    ["98.765.432/0001-10", "01/2025",
     532_000.00, 532_000.00, 0.00, 0.00, 0.0, 0.00, 0.00, 0.00, 0.00, 0.00],
    ["98.765.432/0001-10", "02/2025",
     280_000.00, 280_000.00, 0.00, 0.00, 0.0, 0.00, 0.00, 0.00, 0.00, 0.00],
    ["98.765.432/0001-10", "03/2025",
     190_000.00, 0.00, 9_405.00, 180_595.00, 4.95, 8_939.45, 23_520.00, 0.00, 0.00, 0.00],
    ["98.765.432/0001-10", "04/2025",
     45_000.00, 0.00, 0.00, 45_000.00, 9.9, 4_455.00, 2_540.00, 1_915.00, 1_915.00, 0.00],
]
_save_xlsx_csv("memoria_calculo_CBS", cols_mc, rows_mc)

# ══════════════════════════════════════════════════════════════════════════════
# 8. APURAÇÃO CBS JAN–JUN (consolidado semestral)
# ══════════════════════════════════════════════════════════════════════════════
cols_ap = [
    "CNPJ", "Razão Social", "Período", "Receita Bruta Total (R$)",
    "Base CBS (R$)", "CBS Bruta (R$)", "Créditos (R$)", "CBS Líquida (R$)",
    "Recolhido (R$)", "Saldo (R$)", "Situação",
]
rows_ap = [
    ["12.345.678/0001-90", "Siderúrgica Alfa Ltda", "Jan–Jun/2025",
     2_868_000.00, 2_401_666.32, 238_764.97, 65_880.00, 178_404.97, 178_404.97, 0.00, "Sem diferença"],
    ["98.765.432/0001-10", "Alimentar Beta S.A.", "Jan–Jun/2025",
     1_450_300.00, 225_595.00, 13_394.45, 26_060.00, 0.00, 1_915.00, -1_915.00, "Crédito excedente"],
]
_save_xlsx_csv("apuracao_cbs_jan_jun", cols_ap, rows_ap)

# ══════════════════════════════════════════════════════════════════════════════
# 9. BASE DE CÁLCULO AJUSTADA (inclui ajustes por notas de correção/débito)
# ══════════════════════════════════════════════════════════════════════════════
cols_bc = [
    "CNPJ", "Doc Origem", "Tipo Ajuste", "Competência Original",
    "Valor Original (R$)", "Ajuste (R$)", "Base Ajustada (R$)",
    "CBS Ajuste (R$)", "Motivo",
]
rows_bc = [
    ["12.345.678/0001-90", "NF-e 000001", "Nota Crédito", "01/2025",
     245_000.00, -15_000.00, 230_000.00, -1_485.00, "Devolução parcial mercadorias"],
    ["12.345.678/0001-90", "NF-e 000003", "Nota Débito", "03/2025",
     0.00, 18_000.00, 18_000.00, 1_782.00, "Acréscimo frete internacional reclassificado"],
    ["98.765.432/0001-10", "NF-e 000003", "Nota Crédito", "03/2025",
     190_000.00, -8_000.00, 182_000.00, -396.00, "Abatimento serviço saúde — acerto contratual"],
]
_save_xlsx_csv("base_calculo_ajustada", cols_bc, rows_bc)

# ══════════════════════════════════════════════════════════════════════════════
# 10. DIVERGÊNCIAS IDENTIFICADAS (confronto sistemas RFB)
# ══════════════════════════════════════════════════════════════════════════════
cols_div = [
    "CNPJ", "Competência", "Tipo Divergência", "Valor Declarado (R$)",
    "Valor RFB (R$)", "Diferença (R$)", "% Divergência",
    "Classificação Risco", "Ação Recomendada",
]
rows_div = [
    ["12.345.678/0001-90", "06/2025", "Base CBS — combustível",
     65_288.32, 116_622.00, 51_333.68, 44.0, "ALTO",
     "Verificar aplicação redução setorial Art. 39"],
    ["98.765.432/0001-10", "03/2025", "Alíquota serv. saúde",
     9_405.00, 18_810.00, 9_405.00, 50.0, "ALTO",
     "Confirmar enquadramento Art. 44 — alíquota 50%"],
    ["12.345.678/0001-90", "02/2025", "Crédito transporte",
     5_000.00, 11_760.00, 6_760.00, 57.5, "MEDIO",
     "Checar documentação comprobatória do crédito parcial"],
    ["98.765.432/0001-10", "04/2025", "CBS locação",
     4_455.00, 0.00, -4_455.00, 100.0, "CRITICO",
     "RFB não reconhece CBS sobre locação bem móvel — rever Art. 71"],
]
_save_xlsx_csv("divergencias_identificadas", cols_div, rows_div)

# ══════════════════════════════════════════════════════════════════════════════
# 11. ALÍQUOTAS DE REFERÊNCIA CBS (tabela normativa auxiliar)
# ══════════════════════════════════════════════════════════════════════════════
wb_aliq = openpyxl.Workbook()
ws_aliq = wb_aliq.active
ws_aliq.title = "Aliquotas_CBS"
cols_aliq = [
    "Grupo", "Descrição", "Dispositivo LC 214/2025",
    "Alíquota CBS (%)", "Fator Redução", "Alíquota Efetiva (%)", "Vigência",
]
rows_aliq = [
    ["GERAL", "Operações não especificadas", "Art. 9°", 9.9, 0.0, 9.9, "01/01/2026"],
    ["ALIMENTAR", "Alimentos da cesta básica", "Art. 47 + Anexo I", 9.9, 100.0, 0.0, "01/01/2026"],
    ["ALIMENTAR", "Demais alimentos", "Art. 48", 9.9, 40.0, 5.94, "01/01/2026"],
    ["SAUDE", "Medicamentos essenciais", "Art. 44 + Anexo VII", 9.9, 100.0, 0.0, "01/01/2026"],
    ["SAUDE", "Serviços de saúde humana", "Art. 44", 9.9, 50.0, 4.95, "01/01/2026"],
    ["SAUDE", "Planos de saúde", "Art. 44 §2°", 9.9, 60.0, 3.96, "01/01/2026"],
    ["EDUCACAO", "Ed. básica e superior", "Art. 45", 9.9, 60.0, 3.96, "01/01/2026"],
    ["COMBUSTIVEL", "Derivados petróleo (ICMS-ST)", "Art. 39", 9.9, 44.0, 5.544, "01/01/2026"],
    ["TRANSPORTE", "Coletivo urbano e metro", "Art. 41", 9.9, 100.0, 0.0, "01/01/2026"],
    ["FINANCEIRO", "Receitas financeiras PJ", "Art. 31", 9.9, 0.0, 9.9, "01/01/2026"],
    ["IMOBILIARIO", "Alienação imóvel", "Art. 33 §4°", 9.9, 0.0, 9.9, "01/01/2026"],
    ["AGRO", "Produtos agropecuários", "Art. 46 + Anexo II", 9.9, 100.0, 0.0, "01/01/2026"],
    ["EXPORTACAO", "Exportação de bens e serviços", "Art. 12 §1°", 9.9, 100.0, 0.0, "01/01/2026"],
]
_wb_header(ws_aliq, cols_aliq)
for r, row in enumerate(rows_aliq, 2):
    for c, val in enumerate(row, 1):
        ws_aliq.cell(row=r, column=c, value=val)
ws_aliq.freeze_panes = "A2"
wb_aliq.save(OUT / "aliquotas_referencia_cbs.xlsx")
print("  ✓ aliquotas_referencia_cbs.xlsx")

# ══════════════════════════════════════════════════════════════════════════════
# 12. PLANILHA DE VERIFICAÇÃO (dispara chamadas extras Watson/Sherlock)
# ══════════════════════════════════════════════════════════════════════════════
wb_vf = openpyxl.Workbook()
ws_vf = wb_vf.active
ws_vf.title = "Verificacao"
cols_vf = [
    "Item Verificação", "Arquivo Fonte", "Status",
    "Valor Esperado (R$)", "Valor Encontrado (R$)", "Divergência (R$)",
    "Risco", "Responsável Análise", "Observação DVA",
]
rows_vf = [
    ["V-01 — Base CBS empresa A jan", "memoria_calculo_CBS.xlsx", "PENDENTE",
     84_150.00, 84_150.00, 0.00, "BAIXO", "Watson", "Conferir com NF-e emitidas"],
    ["V-02 — Redução setorial combustível", "reducoes_setoriais.xlsx", "PENDENTE",
     65_288.32, 65_288.32, 0.00, "ALTO", "Sherlock", "LC 214 Art. 39 — cálculo fator"],
    ["V-03 — Crédito PIS/COFINS empresa A", "creditos_pis_cofins.xlsx", "PENDENTE",
     65_880.00, 65_880.00, 0.00, "MEDIO", "Watson", "Verificar documentos de suporte"],
    ["V-04 — Isenção alimentos empresa B", "receita_bruta_empresa_B.xlsx", "PENDENTE",
     0.00, 0.00, 0.00, "ALTO", "Sherlock", "Confirmar lista Art. 47 + Anexo I"],
    ["V-05 — CBS locação bem móvel B", "divergencias_identificadas.xlsx", "CRITICO",
     4_455.00, 0.00, 4_455.00, "CRITICO", "Sherlock", "Divergência RFB — Art. 71 controverso"],
    ["V-06 — Ajuste nota crédito jan A", "base_calculo_ajustada.xlsx", "PENDENTE",
     -1_485.00, -1_485.00, 0.00, "BAIXO", "Watson", "Devolução parcial — ok"],
    ["V-07 — Serviços saúde alíquota B", "divergencias_identificadas.xlsx", "PENDENTE",
     9_405.00, 18_810.00, 9_405.00, "ALTO", "Sherlock", "Alíquota 50% vs cheia"],
    ["V-08 — Exportações zeradas empresa A", "apuracao_cbs_jan_jun.xlsx", "PENDENTE",
     0.00, 0.00, 0.00, "BAIXO", "Watson", "Imunidade exportação — verificar docs"],
]
_wb_header(ws_vf, cols_vf)
for r, row in enumerate(rows_vf, 2):
    for c, val in enumerate(row, 1):
        cell = ws_vf.cell(row=r, column=c, value=val)
        if col_idx_risk := (row[6] if len(row) > 6 else ""):
            pass
ws_vf.freeze_panes = "A2"

# Segunda aba — Checklist normativo
ws_ck = wb_vf.create_sheet("Checklist_Normativo")
cols_ck = ["Dispositivo", "Descrição Verificação", "Aplicável", "Atendido", "Observação"]
rows_ck = [
    ["LC 214 Art. 9°", "Alíquota CBS padrão 9,9%", "Sim", "Sim", "Confirmado nas NF-e"],
    ["LC 214 Art. 12 §1°", "Exportação — imunidade CBS", "Sim", "Sim", "NF-e CFOP 7xxx zeradas"],
    ["LC 214 Art. 39", "Redução setorial combustíveis", "Sim", "Pendente", "Aguarda confirmar fator"],
    ["LC 214 Art. 44", "Alíquota reduzida saúde", "Sim", "Pendente", "Divergência identificada V-07"],
    ["LC 214 Art. 47", "Alíquota zero alimentos básicos", "Sim", "Sim", "Empresa B — ok"],
    ["LC 214 Art. 54", "Créditos PIS/COFINS na transição", "Sim", "Sim", "Documentos verificados"],
    ["LC 214 Art. 71", "Locação bem móvel — incidência", "Sim", "Divergente", "RFB vs contribuinte — V-05"],
    ["LC 214 Anexo I", "Lista alimentos alíquota zero", "Sim", "Sim", "NCMs confirmados"],
    ["LC 214 Anexo VII", "Lista medicamentos essenciais", "Sim", "Sim", "NCMs confirmados"],
]
_wb_header(ws_ck, cols_ck)
for r, row in enumerate(rows_ck, 2):
    for c, val in enumerate(row, 1):
        ws_ck.cell(row=r, column=c, value=val)
wb_vf.save(OUT / "Planilha_Verificacao_MOD_SINT_001.xlsx")
print("  ✓ Planilha_Verificacao_MOD_SINT_001.xlsx")

# ══════════════════════════════════════════════════════════════════════════════
# 13. INVENTÁRIO (arquivo sempre analisado — controle do pacote)
# ══════════════════════════════════════════════════════════════════════════════
wb_inv = openpyxl.Workbook()
ws_inv = wb_inv.active
ws_inv.title = "Inventario"
cols_inv = [
    "N°", "Nome do Arquivo", "Tipo", "Categoria",
    "Tamanho Estimado", "Hash SHA-256 (primeiros 16)", "Responsável Envio",
    "Data Envio", "Observação",
]
rows_inv = [
    [1, "receita_bruta_empresa_A.xlsx", "XLSX", "Analisável", "~18 KB", "a1b2c3d4e5f60001", "RFB-COFIS", "2025-06-05", "Receita bruta 6 competências"],
    [2, "receita_bruta_empresa_A.csv", "CSV", "Analisável", "~3 KB", "b2c3d4e5f6a70002", "RFB-COFIS", "2025-06-05", "Espelho CSV"],
    [3, "receita_bruta_empresa_B.xlsx", "XLSX", "Analisável", "~18 KB", "c3d4e5f6a7b80003", "RFB-COFIS", "2025-06-05", "Receita bruta Empresa B"],
    [4, "receita_bruta_empresa_B.csv", "CSV", "Analisável", "~3 KB", "d4e5f6a7b8c90004", "RFB-COFIS", "2025-06-05", "Espelho CSV"],
    [5, "nfe_emitidas_empresa_A.xlsx", "XLSX", "Analisável", "~20 KB", "e5f6a7b8c9d00005", "RFB-SPED", "2025-06-05", "NF-e emitidas — 5 notas"],
    [6, "nfe_emitidas_empresa_A.csv", "CSV", "Analisável", "~4 KB", "f6a7b8c9d0e10006", "RFB-SPED", "2025-06-05", "Espelho CSV"],
    [7, "nfe_emitidas_empresa_B.xlsx", "XLSX", "Analisável", "~20 KB", "a7b8c9d0e1f20007", "RFB-SPED", "2025-06-05", "NF-e emitidas Empresa B"],
    [8, "nfe_emitidas_empresa_B.csv", "CSV", "Analisável", "~3 KB", "b8c9d0e1f2a30008", "RFB-SPED", "2025-06-05", "Espelho CSV"],
    [9, "creditos_pis_cofins.xlsx", "XLSX", "Analisável", "~16 KB", "c9d0e1f2a3b40009", "RFB-COFIS", "2025-06-05", "Créditos acumulados transição"],
    [10, "creditos_pis_cofins.csv", "CSV", "Analisável", "~2 KB", "d0e1f2a3b4c50010", "RFB-COFIS", "2025-06-05", "Espelho CSV"],
    [11, "reducoes_setoriais.xlsx", "XLSX", "Analisável", "~16 KB", "e1f2a3b4c5d60011", "RFB-COFIS", "2025-06-05", "Cálculo fatores de redução"],
    [12, "reducoes_setoriais.csv", "CSV", "Analisável", "~2 KB", "f2a3b4c5d6e70012", "RFB-COFIS", "2025-06-05", "Espelho CSV"],
    [13, "memoria_calculo_CBS.xlsx", "XLSX", "Analisável", "~22 KB", "a3b4c5d6e7f80013", "RFB-COFIS", "2025-06-05", "10 linhas — apuração mensal"],
    [14, "memoria_calculo_CBS.csv", "CSV", "Analisável", "~4 KB", "b4c5d6e7f8a90014", "RFB-COFIS", "2025-06-05", "Espelho CSV"],
    [15, "apuracao_cbs_jan_jun.xlsx", "XLSX", "Analisável", "~16 KB", "c5d6e7f8a9b00015", "RFB-COFIS", "2025-06-05", "Consolidado semestral"],
    [16, "apuracao_cbs_jan_jun.csv", "CSV", "Analisável", "~1 KB", "d6e7f8a9b0c10016", "RFB-COFIS", "2025-06-05", "Espelho CSV"],
    [17, "base_calculo_ajustada.xlsx", "XLSX", "Analisável", "~16 KB", "e7f8a9b0c1d20017", "RFB-COFIS", "2025-06-05", "Ajustes por notas de correção"],
    [18, "base_calculo_ajustada.csv", "CSV", "Analisável", "~1 KB", "f8a9b0c1d2e30018", "RFB-COFIS", "2025-06-05", "Espelho CSV"],
    [19, "divergencias_identificadas.xlsx", "XLSX", "Analisável", "~18 KB", "a9b0c1d2e3f40019", "DVA-CBS", "2025-06-05", "4 divergências — 1 crítica"],
    [20, "divergencias_identificadas.csv", "CSV", "Analisável", "~2 KB", "b0c1d2e3f4a50020", "DVA-CBS", "2025-06-05", "Espelho CSV"],
    [21, "aliquotas_referencia_cbs.xlsx", "XLSX", "Analisável", "~18 KB", "c1d2e3f4a5b60021", "DVA-CBS", "2025-06-05", "Tabela normativa CBS"],
    [22, "Planilha_Verificacao_MOD_SINT_001.xlsx", "XLSX", "Analisável", "~24 KB", "d2e3f4a5b6c70022", "DVA-CBS", "2025-06-05", "Checklist + 2 abas"],
    [23, "consulta_apuracao_cbs.sql", "SQL", "Analisável", "~3 KB", "e3f4a5b6c7d80023", "DVA-CBS", "2025-06-05", "Query de apuração"],
    [24, "notas_metodologicas.txt", "TXT", "Analisável", "~4 KB", "f4a5b6c7d8e90024", "DVA-CBS", "2025-06-05", "Notas contextuais"],
    [25, "notas_fiscais_amostra.txt", "TXT", "Analisável", "~2 KB", "a5b6c7d8e9f00025", "DVA-CBS", "2025-06-05", "Amostra de NFs em texto"],
    [26, "Metodologia_CBS_PF.md", "MD", "Analisável", "~8 KB", "b6c7d8e9f0a10026", "DVA-CBS", "2025-06-05", "Metodologia para Sherlock"],
    [27, "protocolo_recebimento.md", "MD", "Metadados+Análise", "~3 KB", "c7d8e9f0a1b20027", "DVA-CBS", "2025-06-05", "Protocolo de intake"],
    [28, "inventario.xlsx", "XLSX", "Metadados+Análise", "~26 KB", "d8e9f0a1b2c30028", "DVA-CBS", "2025-06-05", "Este inventário"],
]
_wb_header(ws_inv, cols_inv)
for r, row in enumerate(rows_inv, 2):
    for c, val in enumerate(row, 1):
        ws_inv.cell(row=r, column=c, value=val)
ws_inv.freeze_panes = "A2"
wb_inv.save(OUT / "inventario.xlsx")
print("  ✓ inventario.xlsx")

# ══════════════════════════════════════════════════════════════════════════════
# 14. SQL — Consulta de apuração CBS
# ══════════════════════════════════════════════════════════════════════════════
sql_content = textwrap.dedent("""
    -- consulta_apuracao_cbs.sql
    -- DVA-CBS | MOD_SINT_001 — TC 015.848/2025-6
    -- Query de apuração CBS — confronto declarado × esperado
    -- Base de dados: SPED Fiscal + EFD-Contribuições (espelho RFB)

    -- ─────────────────────────────────────────────────────────────────
    -- 1. Receita bruta por CNPJ e competência
    -- ─────────────────────────────────────────────────────────────────
    WITH receita_bruta AS (
        SELECT
            cnpj,
            competencia,
            SUM(valor_liquido)       AS total_receita_bruta,
            SUM(base_calculo_cbs)    AS total_base_cbs_declarada
        FROM efd_contribuicoes.receita_bruta
        WHERE competencia BETWEEN '2025-01' AND '2025-06'
          AND cnpj IN ('12345678000190', '98765432000110')
        GROUP BY cnpj, competencia
    ),

    -- ─────────────────────────────────────────────────────────────────
    -- 2. CBS calculada pela RFB com alíquotas de referência
    -- ─────────────────────────────────────────────────────────────────
    cbs_rfb AS (
        SELECT
            rb.cnpj,
            rb.competencia,
            rb.total_base_cbs_declarada,
            ar.aliquota_efetiva_pct                            AS aliquota_rfb,
            ROUND(rb.total_base_cbs_declarada
                  * ar.aliquota_efetiva_pct / 100.0, 2)        AS cbs_rfb_calculada
        FROM receita_bruta rb
        LEFT JOIN referencia.aliquotas_cbs ar
            ON ar.vigencia <= rb.competencia
           AND ar.cnpj = rb.cnpj
    ),

    -- ─────────────────────────────────────────────────────────────────
    -- 3. CBS declarada pelo contribuinte (EFD)
    -- ─────────────────────────────────────────────────────────────────
    cbs_declarada AS (
        SELECT
            cnpj,
            competencia,
            SUM(cbs_bruta)           AS cbs_bruta_declarada,
            SUM(creditos_cbs)        AS creditos_declarados,
            SUM(cbs_a_recolher)      AS cbs_liquida_declarada,
            SUM(darf_recolhido)      AS cbs_recolhida
        FROM efd_contribuicoes.apuracao_cbs
        WHERE competencia BETWEEN '2025-01' AND '2025-06'
          AND cnpj IN ('12345678000190', '98765432000110')
        GROUP BY cnpj, competencia
    ),

    -- ─────────────────────────────────────────────────────────────────
    -- 4. Confronto e sinalização de divergências
    -- ─────────────────────────────────────────────────────────────────
    divergencias AS (
        SELECT
            r.cnpj,
            r.competencia,
            r.cbs_rfb_calculada,
            d.cbs_bruta_declarada,
            d.cbs_recolhida,
            ROUND(r.cbs_rfb_calculada - d.cbs_bruta_declarada, 2) AS diferenca_bruta,
            CASE
                WHEN ABS(r.cbs_rfb_calculada - d.cbs_bruta_declarada)
                     / NULLIF(r.cbs_rfb_calculada, 0) > 0.40 THEN 'CRITICO'
                WHEN ABS(r.cbs_rfb_calculada - d.cbs_bruta_declarada)
                     / NULLIF(r.cbs_rfb_calculada, 0) > 0.10 THEN 'ALTO'
                WHEN ABS(r.cbs_rfb_calculada - d.cbs_bruta_declarada)
                     / NULLIF(r.cbs_rfb_calculada, 0) > 0.02 THEN 'MEDIO'
                ELSE 'BAIXO'
            END AS nivel_risco
        FROM cbs_rfb r
        LEFT JOIN cbs_declarada d USING (cnpj, competencia)
    )

    -- ─────────────────────────────────────────────────────────────────
    -- 5. Resultado final para análise DVA
    -- ─────────────────────────────────────────────────────────────────
    SELECT
        cnpj,
        competencia,
        cbs_rfb_calculada,
        cbs_bruta_declarada,
        cbs_recolhida,
        diferenca_bruta,
        nivel_risco,
        CASE
            WHEN nivel_risco IN ('CRITICO', 'ALTO')
            THEN 'Requer análise imediata da DVA-CBS'
            WHEN nivel_risco = 'MEDIO'
            THEN 'Verificar documentação de suporte'
            ELSE 'Dentro da margem — validar amostral'
        END AS recomendacao_dva
    FROM divergencias
    ORDER BY
        CASE nivel_risco
            WHEN 'CRITICO' THEN 1
            WHEN 'ALTO'    THEN 2
            WHEN 'MEDIO'   THEN 3
            ELSE 4
        END,
        cnpj,
        competencia;

    -- ─────────────────────────────────────────────────────────────────
    -- NOTA TÉCNICA (DVA-CBS / 2025-06-05):
    --   Esta query confronta EFD-Contribuições com cálculo independente
    --   baseado nas alíquotas LC 214/2025. As divergências identificadas
    --   nas CTEs "divergencias" são o ponto de partida para análise
    --   do Dr. Watson (integridade técnica) e Sherlock (validação
    --   metodológica de enquadramento normativo).
    --   Referência: TC 015.848/2025-6, Acórdão 2833/2025-Plenário.
    -- ─────────────────────────────────────────────────────────────────
""").strip()

(OUT / "consulta_apuracao_cbs.sql").write_text(sql_content, encoding="utf-8")
print("  ✓ consulta_apuracao_cbs.sql")

# ══════════════════════════════════════════════════════════════════════════════
# 15. TXT — Notas metodológicas
# ══════════════════════════════════════════════════════════════════════════════
notas_met = textwrap.dedent("""
    NOTAS METODOLÓGICAS — ANÁLISE CBS MOD_SINT_001
    DVA-CBS | TC 015.848/2025-6 | TCU / SecexContas
    Data: 05/06/2025

    ══════════════════════════════════════════════════════
    1. ESCOPO DA ANÁLISE
    ══════════════════════════════════════════════════════

    Esta análise abrange dois contribuintes selecionados para validação
    piloto do módulo MOD_SINT_001:

      - Empresa A: CNPJ 12.345.678/0001-90 (Siderúrgica Alfa Ltda)
        * Receita bruta: R$ 2.868.000,00 (Jan–Jun/2025)
        * Setor: siderurgia, TI, exportação, combustíveis
        * Risco principal: redução setorial art. 39 (combustíveis)

      - Empresa B: CNPJ 98.765.432/0001-10 (Alimentar Beta S.A.)
        * Receita bruta: R$ 1.450.300,00 (Jan–Jun/2025)
        * Setor: alimentos, saúde, locação
        * Risco principal: incidência CBS locação bem móvel (art. 71)

    ══════════════════════════════════════════════════════
    2. CRITÉRIOS DE SELEÇÃO DE RISCO
    ══════════════════════════════════════════════════════

    Seleção baseada em cruzamento de dados EFD-Contribuições × NF-e:

    a) CRITÉRIO 1 — REDUÇÃO SETORIAL (art. 39): divergência > 40%
       entre CBS declarada e CBS calculada com alíquota cheia.
       → Empresa A, competência 06/2025: R$ 51.333,68 de diferença.

    b) CRITÉRIO 2 — INCIDÊNCIA DUVIDOSA (art. 71): CBS declarada
       sobre locação de bem móvel sem reconhecimento pela RFB.
       → Empresa B, competência 04/2025: R$ 4.455,00 em litígio.

    c) CRITÉRIO 3 — ALÍQUOTA REDUZIDA SAÚDE (art. 44): CBS declarada
       utilizando alíquota 50% sem comprovação de enquadramento.
       → Empresa B, competência 03/2025: R$ 9.405,00 de divergência.

    ══════════════════════════════════════════════════════
    3. FONTES DE DADOS
    ══════════════════════════════════════════════════════

    Todos os dados são oriundos exclusivamente de:
      - EFD-Contribuições (entregues via SPED fiscal até 30/05/2025)
      - Arquivo NF-e (espelho RFB — exportação até 15/05/2025)
      - Declarações DCTF referentes ao mesmo período

    ATENÇÃO: Os dados deste pacote são sintéticos para fins de teste
    do sistema DVA-CBS. Valores e CNPJs são fictícios.

    ══════════════════════════════════════════════════════
    4. LIMITAÇÕES
    ══════════════════════════════════════════════════════

    - A análise não cobre o período de transição PIS/COFINS → CBS
      anterior a 01/01/2026 (início CBS pleno).
    - Créditos da fase de transição (art. 54) foram estimados com
      base na proporção histórica PIS/COFINS — sujeitos a validação
      documental pela DVA.
    - A incidência CBS sobre locação de bem móvel (art. 71) é objeto
      de controvérsia interpretativa ainda não pacificada pela RFB.

    ══════════════════════════════════════════════════════
    5. PRÓXIMOS PASSOS
    ══════════════════════════════════════════════════════

    1. Watson: verificar integridade técnica dos 10 CSVs/XLSXs.
    2. Sherlock: classificar enquadramento normativo dos 3 critérios.
    3. Mycroft: consolidar e gerar apêndice técnico para relatório.
    4. Lestrade: chancela após revisão do output final.

    Ref.: Acórdão 2833/2025-Plenário, TC 015.848/2025-6.
""").strip()

(OUT / "notas_metodologicas.txt").write_text(notas_met, encoding="utf-8")
print("  ✓ notas_metodologicas.txt")

# ══════════════════════════════════════════════════════════════════════════════
# 16. TXT — Amostra de notas fiscais (texto livre)
# ══════════════════════════════════════════════════════════════════════════════
notas_nf = textwrap.dedent("""
    AMOSTRA DE NOTAS FISCAIS — TEXTO EXTRAÍDO DO XML
    MOD_SINT_001 | DVA-CBS | 05/06/2025

    ──────────────────────────────────────────────────────
    NF-e 000001 | 15/01/2025 | Siderúrgica Alfa → Cliente X
    ──────────────────────────────────────────────────────
    CNPJ Emitente : 12.345.678/0001-90
    CNPJ Dest.    : 11.111.111/0001-11
    CFOP          : 5102 (Venda nacional)
    NCM           : 7308.90.90 (Estruturas metálicas)
    Qtd           : 125 toneladas
    Vl Unitário   : R$ 1.960,00/t
    Vl Total      : R$ 245.000,00
    Base CBS      : R$ 245.000,00
    Alíq CBS      : 9,9%
    CBS Destacado : R$ 24.255,00
    Obs           : Entrega fracionada — 3 remessas jan/2025

    ──────────────────────────────────────────────────────
    NF-e 000003 | 22/03/2025 | Siderúrgica Alfa → Exterior
    ──────────────────────────────────────────────────────
    CNPJ Emitente : 12.345.678/0001-90
    CNPJ Dest.    : EXTERIOR (EUA)
    CFOP          : 7101 (Exportação direta)
    NCM           : 7308.90.90
    Vl Total      : R$ 415.000,00
    Base CBS      : R$ 0,00  (imunidade exportação)
    CBS Destacado : R$ 0,00
    Obs           : Despacho DUE 2025-03-20 — AWB 123456789

    ──────────────────────────────────────────────────────
    NF-e 000002 (Empresa B) | 12/02/2025 | Alimentar Beta
    ──────────────────────────────────────────────────────
    CNPJ Emitente : 98.765.432/0001-10
    CNPJ Dest.    : 77.777.777/0001-77 (farmácia)
    CFOP          : 5102
    NCM           : 3003.90.89 (Medicamento genérico)
    Qtd           : 12.400 unidades
    Vl Unitário   : R$ 7,42/un
    Vl Total      : R$ 92.000,00
    Base CBS      : R$ 0,00  (isenção — LC 214 Anexo VII)
    CBS Destacado : R$ 0,00
    Obs           : Medicamento fitoterápico — enquadramento Anvisa confirmado

    ──────────────────────────────────────────────────────
    NF-e 000004 (Empresa B) | 25/04/2025 | Locação
    ──────────────────────────────────────────────────────
    CNPJ Emitente : 98.765.432/0001-10
    CNPJ Dest.    : 99.999.999/0001-99
    CFOP          : 0000 (Outros)
    Descrição     : Locação de empilhadeira elétrica 3t
    Período       : Abril/2025
    Vl Total      : R$ 45.000,00
    Base CBS      : R$ 45.000,00  (contribuinte tributa)
    Alíq CBS      : 9,9%
    CBS Destacado : R$ 4.455,00
    CONTROVÉRSIA  : RFB questiona incidência CBS sobre locação (art. 71 LC 214)
                    Aguarda solução de consulta COSIT n° 45/2025.
""").strip()

(OUT / "notas_fiscais_amostra.txt").write_text(notas_nf, encoding="utf-8")
print("  ✓ notas_fiscais_amostra.txt")

# ══════════════════════════════════════════════════════════════════════════════
# 17. METODOLOGIA CBS (Markdown — contexto para Sherlock)
# ══════════════════════════════════════════════════════════════════════════════
met_md = textwrap.dedent("""
    # Metodologia de Apuração CBS — DVA-CBS
    **TC 015.848/2025-6 | LC 214/2025 | Acórdão 2833/2025-Plenário**

    ## 1. Estrutura da CBS

    A Contribuição sobre Bens e Serviços (CBS) é tributo federal criado pela
    LC 214/2025 como componente do IBS-CBS (Reforma Tributária — PEC 45/2019).
    Substitui PIS e COFINS a partir de 01/01/2026 (fase plena).

    **Alíquota padrão:** 9,9% sobre a base de cálculo (art. 9°, LC 214/2025).

    ## 2. Base de Cálculo

    A base de cálculo é o valor da operação, incluindo:
    - Frete e seguro cobrado do adquirente
    - Descontos condicionais
    - Receitas financeiras da PJ (art. 31)

    Excluem-se da base:
    - Exportações (imunidade constitucional — art. 12 §1°)
    - Transferências entre estabelecimentos do mesmo contribuinte
    - ICMS-ST (desde que destacado na NF-e)

    ## 3. Regimes Especiais e Alíquotas Diferenciadas

    | Setor | Dispositivo | Alíq. Efetiva | Condição |
    |-------|-------------|---------------|----------|
    | Alimentos básicos (Anexo I) | Art. 47 | 0% | Lista NCM taxativa |
    | Demais alimentos | Art. 48 | 5,94% | Fator 0,4 sobre 9,9% |
    | Medicamentos essenciais (Anexo VII) | Art. 44 | 0% | Registro Anvisa |
    | Serviços de saúde | Art. 44 | 4,95% | Fator 0,5 |
    | Educação básica/superior | Art. 45 | 3,96% | Fator 0,4 |
    | Combustíveis derivados petróleo | Art. 39 | 5,544% | Fator 0,44 |
    | Transporte coletivo urbano | Art. 41 | 0% | Concessão pública |
    | Exportação bens/serviços | Art. 12 §1° | 0% | Imunidade constitucional |

    ## 4. Créditos da Fase de Transição

    O art. 54 da LC 214/2025 permite o aproveitamento dos saldos credores
    de PIS e COFINS acumulados até 31/12/2025 na forma de crédito CBS:

    - Crédito integral para insumos (§2°)
    - Crédito para energia elétrica (§3° I)
    - Crédito para serviços de transporte de insumos (§3° II)
    - Crédito para arrendamento de máquinas (§4°)

    Método de conversão: saldo PIS/COFINS × fator 1,00 (1:1 na fase de transição).

    ## 5. Pontos Controvertidos — Alertas DVA

    ### 5.1 Locação de Bem Móvel (art. 71)
    A LC 214/2025 inclui locação de bens móveis na hipótese de incidência CBS,
    mas a Receita Federal, pela Solução de Consulta COSIT (pendente), sinaliza
    que a incidência depende da caracterização como "operação onerosa com bem
    móvel". Risco de autuação para contribuintes que não tributam.

    ### 5.2 Redução Setorial — Combustíveis (art. 39)
    O fator de redução de 44% (resultado: alíquota 5,544%) aplica-se apenas
    a derivados de petróleo sujeitos a ICMS-ST. Biocombustíveis e misturas
    têm tratamento diferenciado — exige verificação NCM × Lista art. 39.

    ### 5.3 Serviços de Saúde (art. 44)
    A redução de 50% aplica-se a serviços prestados por profissionais de saúde
    habilitados pelo conselho de classe competente. Serviços administrativos
    de hospitais não se enquadram — risco de subapuração.

    ## 6. Estrutura dos Arquivos do Pacote MOD_SINT_001

    Os arquivos deste pacote cobrem:
    1. Receita bruta declarada (EFD-Contribuições)
    2. NF-e emitidas (confronto com EFD)
    3. Créditos de transição PIS/COFINS → CBS
    4. Reduções setoriais aplicadas
    5. Memória de cálculo mensal
    6. Apuração semestral consolidada
    7. Ajustes por notas de correção/débito
    8. Divergências identificadas no cruzamento
    9. Tabela de alíquotas de referência

    **Foco da análise DVA:** verificar se as reduções de base e alíquota
    aplicadas pelos contribuintes estão aderentes à LC 214/2025 e
    Regulamento CBS (em edição pela RFB).

    ---
    *Documento de uso interno restrito — DVA-CBS / TCU / SecexContas*
    *TC 015.848/2025-6 | Acórdão 2833/2025-Plenário*
""").strip()

(OUT / "Metodologia_CBS_PF.md").write_text(met_md, encoding="utf-8")
print("  ✓ Metodologia_CBS_PF.md")

# ══════════════════════════════════════════════════════════════════════════════
# 18. PROTOCOLO DE RECEBIMENTO (arquivo always-analyzed)
# ══════════════════════════════════════════════════════════════════════════════
proto_md = textwrap.dedent("""
    # Protocolo de Recebimento — MOD_SINT_001
    **DVA-CBS | TC 015.848/2025-6 | TCU / SecexContas**

    | Campo | Valor |
    |-------|-------|
    | Módulo | MOD_SINT_001 |
    | Atividade | 1 — Validação CBS |
    | Data/hora recebimento | 2025-06-05T14:30:00Z |
    | Responsável recebimento | Insp. Lestrade / DVA-CBS |
    | Total de arquivos | 28 |
    | Arquivos para análise | 26 |
    | Hash de integridade do pacote | `a1b2c3d4e5f6789012345678` |

    ## Arquivos Recebidos

    | # | Arquivo | Tipo | Status |
    |---|---------|------|--------|
    | 1 | receita_bruta_empresa_A.xlsx | XLSX | ✓ Íntegro |
    | 2 | receita_bruta_empresa_A.csv | CSV | ✓ Íntegro |
    | 3 | receita_bruta_empresa_B.xlsx | XLSX | ✓ Íntegro |
    | 4 | receita_bruta_empresa_B.csv | CSV | ✓ Íntegro |
    | 5 | nfe_emitidas_empresa_A.xlsx | XLSX | ✓ Íntegro |
    | 6 | nfe_emitidas_empresa_A.csv | CSV | ✓ Íntegro |
    | 7 | nfe_emitidas_empresa_B.xlsx | XLSX | ✓ Íntegro |
    | 8 | nfe_emitidas_empresa_B.csv | CSV | ✓ Íntegro |
    | 9 | creditos_pis_cofins.xlsx | XLSX | ✓ Íntegro |
    | 10 | creditos_pis_cofins.csv | CSV | ✓ Íntegro |
    | 11 | reducoes_setoriais.xlsx | XLSX | ✓ Íntegro |
    | 12 | reducoes_setoriais.csv | CSV | ✓ Íntegro |
    | 13 | memoria_calculo_CBS.xlsx | XLSX | ✓ Íntegro |
    | 14 | memoria_calculo_CBS.csv | CSV | ✓ Íntegro |
    | 15 | apuracao_cbs_jan_jun.xlsx | XLSX | ✓ Íntegro |
    | 16 | apuracao_cbs_jan_jun.csv | CSV | ✓ Íntegro |
    | 17 | base_calculo_ajustada.xlsx | XLSX | ✓ Íntegro |
    | 18 | base_calculo_ajustada.csv | CSV | ✓ Íntegro |
    | 19 | divergencias_identificadas.xlsx | XLSX | ✓ Íntegro |
    | 20 | divergencias_identificadas.csv | CSV | ✓ Íntegro |
    | 21 | aliquotas_referencia_cbs.xlsx | XLSX | ✓ Íntegro |
    | 22 | Planilha_Verificacao_MOD_SINT_001.xlsx | XLSX | ✓ Íntegro |
    | 23 | consulta_apuracao_cbs.sql | SQL | ✓ Íntegro |
    | 24 | notas_metodologicas.txt | TXT | ✓ Íntegro |
    | 25 | notas_fiscais_amostra.txt | TXT | ✓ Íntegro |
    | 26 | Metodologia_CBS_PF.md | MD | ✓ Íntegro |
    | 27 | protocolo_recebimento.md | MD | (este arquivo) |
    | 28 | inventario.xlsx | XLSX | ✓ Íntegro |

    ## Declaração de Integridade

    Todos os arquivos listados foram recebidos sem adulteração confirmada
    via hash SHA-256. Os originais são mantidos intactos em conformidade
    com o Art. 13 do regimento interno DVA-CBS (Projeto Diógenes — Motor de Start).

    **Assinatura Lestrade:** ████████████ [dados fiscais — uso interno]

    ---
    *Uso interno restrito — DVA-CBS / TCU / SecexContas*
""").strip()

(OUT / "protocolo_recebimento.md").write_text(proto_md, encoding="utf-8")
print("  ✓ protocolo_recebimento.md")

# ══════════════════════════════════════════════════════════════════════════════
# Resumo final
# ══════════════════════════════════════════════════════════════════════════════
arquivos = sorted(OUT.iterdir())
print(f"\n{'─'*60}")
print(f"Massa de teste criada em: {OUT}")
print(f"Total de arquivos: {len(arquivos)}")
print(f"{'─'*60}")
for f in arquivos:
    size = f.stat().st_size
    print(f"  {f.name:<52} {size:>7,} bytes")
print(f"{'─'*60}")
print("Pronto! Execute o teste com:")
print("  diogenes autorun --module MOD_SINT_001 --activity 1 --auto-seal")
